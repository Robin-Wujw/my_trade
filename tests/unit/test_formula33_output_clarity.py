import pandas as pd
from openpyxl import load_workbook

from stock_research.pipelines import formula33


def test_fetch_progress_explains_outcomes_and_remaining_work():
    counts = {
        "processed": 0,
        "succeeded": 0,
        "cache_hits": 0,
        "skipped": 0,
        "failed": 0,
        "signals": 0,
    }
    formula33.update_fetch_progress(
        counts,
        [
            {"signal_type": "XG"},
            {
                "signal_type": "STATUS",
                "observation_status": "tradable",
                "data_origin": "complete_file_cache",
            },
        ],
    )
    formula33.update_fetch_progress(
        counts,
        [{"signal_type": "STATUS", "observation_status": "data_unavailable"}],
    )
    formula33.update_fetch_progress(counts, [])

    assert formula33.format_fetch_progress(counts, 10) == (
        "行情进度 3/10 | 成功 1 | 完整缓存 1 | 跳过 1 | "
        "失败 1 | 剩余 7 | BASE/XG记录 1"
    )


def test_formula_workbook_starts_with_plain_language_guide(monkeypatch, tmp_path):
    monkeypatch.setattr(formula33, "OUTPUT_DIR", str(tmp_path))
    summary, hits = formula33.build_sample(5)

    xlsx_path, csv_path = formula33.save_workbook(summary, hits, sample=True)

    workbook = load_workbook(xlsx_path, read_only=True)
    assert workbook.sheetnames[:4] == [
        "先看这里",
        "33公式日统计",
        "横向统计",
        "命中股票",
    ]
    guide = workbook["先看这里"]
    assert guide["A1"].value == "阅读顺序"
    assert "正式名单" in guide["B1"].value
    summary_sheet = workbook["33公式日统计"]
    assert summary_sheet["A1"].value == "日期"
    assert summary_sheet["B1"].value == "当日BASE数量"
    assert "近21日正式名单" in workbook.sheetnames

    machine_frame = pd.read_csv(csv_path)
    assert {"date", "base_count", "count"}.issubset(machine_frame.columns)
