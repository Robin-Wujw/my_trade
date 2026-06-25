# -*- coding: utf-8 -*-
"""
33 formula market-structure statistics.

Formula:
KD := (CLOSE - LLV(LOW, 9)) / (HHV(HIGH, 9) - LLV(LOW, 9)) * 100;
K := SMA(KD, 3, 1);
WR1 := 100 * (HHV(HIGH, 10) - CLOSE) / (HHV(HIGH, 10) - LLV(LOW, 10));
WR2 := 100 * (HHV(HIGH, 20) - CLOSE) / (HHV(HIGH, 20) - LLV(LOW, 20));
KD80 := K > 80;
WR3 := WR1 < 20 AND WR2 < 20;
RSI70 := SMA(MAX(CLOSE - REF(CLOSE, 1), 0), 9, 1)
    / SMA(ABS(CLOSE - REF(CLOSE, 1)), 9, 1) * 100 > 70;
MKT_CAP := FINANCE(40) / 10000 > 100;
LIST_DAYS := FINANCE(42) > 300;
BASE := KD80 AND WR3 AND RSI70 AND MKT_CAP AND LIST_DAYS;
XG: COUNT(BASE, 5) = 5;

The script records the number of Shanghai/Shenzhen A shares matching XG on
the latest N trading days. Rising for 3/5 days means initial/confirmed
structure improvement; falling for 3/5 days means initial/confirmed weakness.
"""
import argparse
from concurrent.futures import ThreadPoolExecutor, as_completed
import json
import multiprocessing
import os
import random
import time
from datetime import datetime

import akshare as ak
import baostock as bs
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from trade_utils import get_project_path


if not hasattr(pd.DataFrame, "append"):
    def _df_append(self, other, ignore_index=False, **kwargs):
        return pd.concat([self, other], ignore_index=ignore_index)

    pd.DataFrame.append = _df_append


OUTPUT_DIR = get_project_path("板块观察")
ADJUST_FLAG_QFQ = "2"
SHARE_CACHE_FILE = get_project_path(".cache/formula33_share_capital.json")
KLINE_CACHE_DIR = get_project_path(".cache/formula33_kline")
REQUEST_RETRY_ERRORS = (BrokenPipeError, ConnectionError, TimeoutError, OSError)


def parse_args():
    parser = argparse.ArgumentParser(description="33公式沪深A股市场结构统计")
    parser.add_argument("--lookback", type=int, default=21, help="统计最近N个交易日")
    parser.add_argument("--start-date", default="", help="统计起始交易日，格式 YYYY-MM-DD；传入后按日期区间统计")
    parser.add_argument("--end-date", default="", help="统计截止交易日，格式 YYYY-MM-DD；留空则使用当前日期")
    parser.add_argument("--history-days", type=int, default=90, help="为指标计算额外拉取的自然日长度")
    parser.add_argument("--workers", type=int, default=4)
    parser.add_argument("--sleep", type=float, default=0.0)
    parser.add_argument("--retries", type=int, default=4, help="单只股票行情请求失败重试次数")
    parser.add_argument("--retry-delay", type=float, default=1.5, help="请求失败后的退避基准秒数")
    parser.add_argument("--capital-workers", type=int, default=2, help="沪市股本结构补充并发数")
    parser.add_argument("--capital-sleep", type=float, default=0.08, help="沪市股本结构补充每个请求前等待秒数")
    parser.add_argument("--offset", type=int, default=0, help="分段续跑用，从股票池第N只开始")
    parser.add_argument("--limit", type=int, default=0, help="调试用，只处理前N只")
    parser.add_argument("--exclude-codes", default="", help="逗号分隔的股票代码，按截图/外部股票池复核时可排除")
    parser.add_argument("--maxtasksperchild", type=int, default=200, help="多进程模式下每个worker处理多少任务后重启")
    parser.add_argument("--price-source", choices=["baostock", "akshare"], default="akshare", help="前复权K线来源")
    parser.add_argument("--min-mktcap", type=float, default=100.0, help="最低总市值，单位亿元")
    parser.add_argument("--min-list-days", type=int, default=300, help="最低上市天数")
    parser.add_argument("--debug-filters", action="store_true", help="打印最近交易日各条件逐步通过数量")
    parser.add_argument("--require-end-trade", action="store_true", help="只保留截止日有K线的股票，用于复核当前交易列表并排除停牌票")
    parser.add_argument(
        "--market-cap-source",
        choices=["auto", "tushare", "akshare", "akshare-capital", "none"],
        default="auto",
        help="总市值来源；none 仅用于临时复核技术指标数量，会跳过 FINANCE(40) 过滤",
    )
    parser.add_argument("--sample", action="store_true", help="生成离线样例，不访问网络")
    return parser.parse_args()


def call_with_backoff(func, label, retries=4, retry_delay=1.5):
    last_exc = None
    for attempt in range(1, max(1, retries) + 1):
        try:
            return func()
        except Exception as exc:
            last_exc = exc
            if attempt >= retries:
                break
            wait = retry_delay * attempt + random.uniform(0, retry_delay)
            print(f"{label} 请求失败: {exc} | 第 {attempt}/{retries} 次，{wait:.1f}s 后重试")
            time.sleep(wait)
    raise last_exc


def to_bs_code(raw_code):
    code = str(raw_code).strip()
    if "." in code:
        return code
    low = code.lower()
    if low.startswith("sh") and len(code) >= 8:
        return f"sh.{code[-6:]}"
    if low.startswith("sz") and len(code) >= 8:
        return f"sz.{code[-6:]}"
    if low.startswith("bj") and len(code) >= 8:
        return f"bj.{code[-6:]}"
    code = code.zfill(6)
    if code.startswith(("6", "9")):
        return f"sh.{code}"
    return f"sz.{code}"


def pure_code(bs_code):
    return str(bs_code).split(".")[-1]


def normalize_six_digit_code(raw_code):
    text = str(raw_code).strip()
    if text.endswith(".0"):
        text = text[:-2]
    digits = "".join(ch for ch in text if ch.isdigit())
    return digits[-6:].zfill(6) if digits else ""


def get_trade_dates(lookback, extra_days):
    end = pd.Timestamp.today().strftime("%Y-%m-%d")
    start = (pd.Timestamp.today() - pd.DateOffset(days=extra_days)).strftime("%Y-%m-%d")
    rs = bs.query_trade_dates(start_date=start, end_date=end)
    df = rs.get_data()
    if rs.error_code != "0" or df.empty:
        return []
    df.columns = rs.fields
    df = df[df["is_trading_day"] == "1"]
    return df["calendar_date"].tail(lookback).tolist()


def get_trade_dates_akshare(lookback, extra_days):
    df = ak.tool_trade_date_hist_sina()
    if df is None or df.empty:
        return []
    df = df.copy()
    df["trade_date"] = pd.to_datetime(df["trade_date"], errors="coerce")
    end = pd.Timestamp.today().normalize()
    start = end - pd.DateOffset(days=extra_days)
    df = df[(df["trade_date"] >= start) & (df["trade_date"] <= end)].dropna(subset=["trade_date"])
    return df["trade_date"].dt.strftime("%Y-%m-%d").tail(lookback).tolist()


def select_trade_dates(trade_dates, start_date, end_date, lookback):
    if start_date:
        start = pd.to_datetime(start_date, errors="coerce")
        if pd.isna(start):
            raise SystemExit(f"--start-date 格式错误: {start_date}")
        start_text = start.strftime("%Y-%m-%d")
        trade_dates = [date for date in trade_dates if date >= start_text]
    if end_date:
        cap = pd.to_datetime(end_date, errors="coerce")
        if pd.isna(cap):
            raise SystemExit(f"--end-date 格式错误: {end_date}")
        cap_text = cap.strftime("%Y-%m-%d")
        trade_dates = [date for date in trade_dates if date <= cap_text]
    if start_date:
        return trade_dates
    return trade_dates[-lookback:]


def parse_code_set(value):
    codes = set()
    for item in str(value or "").replace("，", ",").split(","):
        item = item.strip()
        if item:
            codes.add(to_bs_code(item))
    return codes


def get_universe(latest_date):
    rs = bs.query_all_stock(day=latest_date)
    df = rs.get_data()
    if rs.error_code != "0" or df.empty:
        return pd.DataFrame()
    df.columns = rs.fields
    mask = (
        df["code"].str.startswith("sh.60")
        | df["code"].str.startswith("sh.68")
        | df["code"].str.startswith("sz.00")
        | df["code"].str.startswith("sz.30")
    )
    df = df[mask].copy()
    if "tradeStatus" in df.columns:
        df = df[df["tradeStatus"] != "0"]
    if "code_name" in df.columns:
        df = df[~df["code_name"].astype(str).str.contains("ST", na=False)]
    return df[["code", "code_name"]].drop_duplicates("code")


def get_universe_with_fallback(trade_dates):
    for date in reversed(trade_dates):
        df = get_universe(date)
        if not df.empty:
            return date, df
    return "", pd.DataFrame()


def get_universe_akshare():
    df = ak.stock_info_a_code_name()
    if df is None or df.empty:
        return pd.DataFrame()
    df = df.rename(columns={"code": "raw_code", "name": "code_name"})
    df["code"] = df["raw_code"].map(to_bs_code)
    mask = (
        df["code"].str.startswith("sh.60")
        | df["code"].str.startswith("sh.68")
        | df["code"].str.startswith("sz.00")
        | df["code"].str.startswith("sz.30")
    )
    df = df[mask].copy()
    df = df[~df["code_name"].astype(str).str.contains("ST|退", na=False)]
    return df[["code", "code_name"]].drop_duplicates("code")


def load_stock_basic():
    try:
        rs = bs.query_stock_basic()
        df = rs.get_data()
    except Exception:
        return pd.DataFrame()
    if df is None or df.empty:
        return pd.DataFrame()
    df.columns = rs.fields
    return df


def load_stock_basic_akshare():
    rows = []
    for symbol in ["\u4e3b\u677fA\u80a1", "\u79d1\u521b\u677f"]:
        try:
            sh = ak.stock_info_sh_name_code(symbol=symbol)
            for _, row in sh.iterrows():
                rows.append({
                    "code": f"sh.{str(row.get('证券代码')).zfill(6)}",
                    "ipoDate": row.get("上市日期"),
                })
        except Exception:
            continue
    try:
        sz = ak.stock_info_sz_name_code(symbol="\u0041\u80a1\u5217\u8868")
        for _, row in sz.iterrows():
            rows.append({
                "code": f"sz.{str(row.get('A股代码')).zfill(6)}",
                "ipoDate": row.get("A股上市日期"),
            })
    except Exception:
        pass
    return pd.DataFrame(rows)


def get_tushare_token():
    token = os.environ.get("TUSHARE_TOKEN", "").strip()
    if token:
        return token
    token_file = os.environ.get("TUSHARE_TOKEN_FILE", get_project_path(".tushare_token"))
    try:
        with open(token_file, "r", encoding="utf-8") as fh:
            return fh.read().strip()
    except OSError:
        return ""


def load_market_caps_from_tushare(trade_date):
    try:
        import tushare as ts
    except ImportError as exc:
        raise RuntimeError("未安装 tushare，请先 pip install tushare") from exc
    token = get_tushare_token()
    if not token:
        raise RuntimeError("未配置 TUSHARE_TOKEN 或 .tushare_token")
    pro = ts.pro_api(token)
    ts_date = str(trade_date).replace("-", "")
    df = pro.daily_basic(trade_date=ts_date, fields="ts_code,total_mv")
    if df is None or df.empty:
        raise RuntimeError(f"tushare daily_basic 在 {trade_date} 无数据")
    caps = {}
    for _, row in df.iterrows():
        ts_code = str(row.get("ts_code", ""))
        code = ts_code.split(".")[0]
        if ts_code.endswith(".SH"):
            bs_code = f"sh.{code}"
        elif ts_code.endswith(".SZ"):
            bs_code = f"sz.{code}"
        else:
            bs_code = to_bs_code(code)
        total_mv_wan = pd.to_numeric(row.get("total_mv"), errors="coerce")
        if pd.notna(total_mv_wan):
            caps[bs_code] = float(total_mv_wan) / 10000.0
    return caps


def load_market_caps_from_akshare():
    try:
        df = ak.stock_zh_a_spot_em()
    except Exception as exc:
        raise RuntimeError(f"无法获取东方财富A股总市值：{exc}")
    if df is None or df.empty:
        return {}
    code_col = "代码" if "代码" in df.columns else None
    cap_col = "总市值" if "总市值" in df.columns else None
    if not code_col or not cap_col:
        raise RuntimeError("东方财富行情表缺少 代码/总市值 字段")
    caps = {}
    for _, row in df.iterrows():
        cap = pd.to_numeric(row.get(cap_col), errors="coerce")
        if pd.isna(cap):
            continue
        caps[to_bs_code(row.get(code_col))] = float(cap) / 100000000.0
    return caps


def load_share_cache():
    try:
        if os.path.exists(SHARE_CACHE_FILE):
            with open(SHARE_CACHE_FILE, "r", encoding="utf-8") as fh:
                return json.load(fh)
    except (OSError, json.JSONDecodeError):
        return {}
    return {}


def save_share_cache(cache):
    try:
        os.makedirs(os.path.dirname(SHARE_CACHE_FILE), exist_ok=True)
        with open(SHARE_CACHE_FILE, "w", encoding="utf-8") as fh:
            json.dump(cache, fh, ensure_ascii=False, indent=2)
    except OSError:
        pass


def parse_share(value):
    text = str(value).replace(",", "").strip()
    val = pd.to_numeric(text, errors="coerce")
    return None if pd.isna(val) else float(val)


def em_symbol_from_bs(code):
    pure = pure_code(code)
    if str(code).startswith("sh."):
        return f"{pure}.SH"
    if str(code).startswith("sz."):
        return f"{pure}.SZ"
    return pure


def fetch_total_share_from_gbjg(task):
    code, sleep, retries, retry_delay = task
    if sleep > 0:
        time.sleep(sleep)
    try:
        gb = call_with_backoff(
            lambda: ak.stock_zh_a_gbjg_em(symbol=em_symbol_from_bs(code)),
            f"{code} 股本结构",
            retries=retries,
            retry_delay=retry_delay,
        )
        if gb is None or gb.empty or "总股本" not in gb.columns:
            return code, None
        gb = gb.copy()
        gb["变更日期_dt"] = pd.to_datetime(gb["变更日期"], errors="coerce")
        gb = gb.sort_values("变更日期_dt", ascending=False)
        return code, parse_share(gb.iloc[0].get("总股本"))
    except Exception:
        return code, None


def load_market_caps_from_akshare_capital(universe, capital_workers=2, capital_sleep=0.08, retries=4, retry_delay=1.5):
    spot = ak.stock_zh_a_spot()
    if spot is None or spot.empty:
        raise RuntimeError("akshare stock_zh_a_spot 无数据，无法取得当前价格")
    code_col = "代码"
    price_col = "最新价"
    price_map = {}
    for _, row in spot.iterrows():
        price = pd.to_numeric(row.get(price_col), errors="coerce")
        if pd.notna(price) and price > 0:
            price_map[to_bs_code(row.get(code_col))] = float(price)

    share_map = {}
    try:
        sz = ak.stock_info_sz_name_code(symbol="\u0041\u80a1\u5217\u8868")
        for _, row in sz.iterrows():
            raw_code = normalize_six_digit_code(row.get("A股代码"))
            if not raw_code:
                continue
            code = f"sz.{raw_code}"
            share = parse_share(row.get("A股总股本"))
            if share:
                share_map[code] = share
    except Exception as exc:
        print(f"深市总股本读取失败: {exc}")

    cache = load_share_cache()
    missing_sh = []
    for _, row in universe.iterrows():
        code = row["code"]
        if code in share_map:
            continue
        cached = cache.get(code)
        if cached and cached.get("total_share"):
            share_map[code] = float(cached["total_share"])
            continue
        if not code.startswith("sh."):
            continue
        missing_sh.append(code)
    if missing_sh:
        workers = max(1, capital_workers)
        print(f"沪市总股本需补充 {len(missing_sh)} 只，使用 akshare 股本结构读取，并发 {workers}")
        with ThreadPoolExecutor(max_workers=workers) as executor:
            futures = [
                executor.submit(fetch_total_share_from_gbjg, (code, capital_sleep, retries, retry_delay))
                for code in missing_sh
            ]
            for idx, future in enumerate(as_completed(futures), start=1):
                code, share = future.result()
                if share:
                    share_map[code] = share
                    cache[code] = {"total_share": share, "updated_at": datetime.now().strftime("%Y-%m-%d")}
                if idx % 200 == 0:
                    print(f"  沪市总股本进度 {idx}/{len(missing_sh)}")
    save_share_cache(cache)

    caps = {}
    for code, share in share_map.items():
        price = price_map.get(code)
        if price:
            caps[code] = share * price / 100000000.0
    if not caps:
        raise RuntimeError("akshare 股本结构法未生成有效总市值")
    return caps


def load_market_caps(source, trade_date, universe=None, capital_workers=2, capital_sleep=0.08, retries=4, retry_delay=1.5):
    if source == "none":
        print("已按 --market-cap-source none 跳过 FINANCE(40) 市值过滤，仅用于技术条件复核。")
        return {}, "none"

    errors = []
    sources = ["tushare", "akshare", "akshare-capital"] if source == "auto" else [source]
    for item in sources:
        try:
            if item == "tushare":
                caps = load_market_caps_from_tushare(trade_date)
            elif item == "akshare":
                caps = load_market_caps_from_akshare()
            elif item == "akshare-capital":
                if universe is None:
                    raise RuntimeError("akshare-capital 需要股票池 universe")
                caps = load_market_caps_from_akshare_capital(
                    universe,
                    capital_workers=capital_workers,
                    capital_sleep=capital_sleep,
                    retries=retries,
                    retry_delay=retry_delay,
                )
            else:
                continue
            if caps:
                print(f"总市值来源: {item}，记录数 {len(caps)}")
                return caps, item
        except Exception as exc:
            errors.append(f"{item}: {exc}")
    raise RuntimeError("无法获取总市值数据；" + " | ".join(errors))


def tdx_sma(series, n, m=1):
    """Tongdaxin SMA(X,N,M): Y=(M*X+(N-M)*Y')/N."""
    prev = np.nan
    values = []
    for raw in pd.to_numeric(series, errors="coerce"):
        if pd.isna(raw):
            values.append(np.nan)
            continue
        if pd.isna(prev):
            prev = raw
        else:
            prev = (m * raw + (n - m) * prev) / n
        values.append(prev)
    return pd.Series(values, index=series.index)


def calc_kdj_k(df, n=9):
    low_n = df["low"].rolling(n, min_periods=n).min()
    high_n = df["high"].rolling(n, min_periods=n).max()
    kd = (df["close"] - low_n) / (high_n - low_n).replace(0, np.nan) * 100
    return tdx_sma(kd, 3, 1)


def calc_wr(df, n):
    high_n = df["high"].rolling(n, min_periods=n).max()
    low_n = df["low"].rolling(n, min_periods=n).min()
    return (high_n - df["close"]) / (high_n - low_n).replace(0, np.nan) * 100


def calc_rsi(series, n=9):
    diff = series.diff()
    up = diff.clip(lower=0)
    abs_diff = diff.abs()
    return tdx_sma(up, n, 1) / tdx_sma(abs_diff, n, 1).replace(0, np.nan) * 100


def init_worker():
    bs.login()


def kline_cache_path(source, code):
    safe_code = str(code).replace(".", "_")
    return os.path.join(KLINE_CACHE_DIR, source, f"{safe_code}.csv")


def normalize_kline_df(df):
    if df is None or df.empty:
        return pd.DataFrame()
    df = df.copy()
    keep_cols = ["date", "code", "open", "high", "low", "close", "volume", "tradestatus"]
    for col in keep_cols:
        if col not in df.columns:
            df[col] = np.nan
    df = df[keep_cols]
    df["date"] = pd.to_datetime(df["date"], errors="coerce").dt.strftime("%Y-%m-%d")
    for col in ["open", "high", "low", "close", "volume"]:
        df[col] = pd.to_numeric(df[col], errors="coerce")
    df = df.dropna(subset=["date", "high", "low", "close"]).sort_values("date")
    return df.drop_duplicates(subset=["date"], keep="last").reset_index(drop=True)


def load_cached_kline(source, code):
    path = kline_cache_path(source, code)
    try:
        if os.path.exists(path):
            return normalize_kline_df(pd.read_csv(path, dtype={"code": str}))
    except Exception as exc:
        print(f"{code} K线缓存读取失败: {exc}")
    return pd.DataFrame()


def save_cached_kline(source, code, df):
    if df is None or df.empty:
        return
    path = kline_cache_path(source, code)
    try:
        os.makedirs(os.path.dirname(path), exist_ok=True)
        tmp_path = f"{path}.{os.getpid()}.tmp"
        normalize_kline_df(df).to_csv(tmp_path, index=False, encoding="utf-8-sig")
        os.replace(tmp_path, path)
    except OSError as exc:
        print(f"{code} K线缓存保存失败: {exc}")


def filter_kline_range(df, start_date, end_date):
    if df is None or df.empty:
        return pd.DataFrame()
    df = normalize_kline_df(df)
    mask = (df["date"] >= start_date) & (df["date"] <= end_date)
    return df[mask].copy().reset_index(drop=True)


def load_kline_with_cache(source, code, start_date, end_date, retries=4, retry_delay=1.5):
    cached = load_cached_kline(source, code)
    if not cached.empty:
        cached_min = cached["date"].min()
        cached_max = cached["date"].max()
        if cached_min <= start_date and cached_max >= end_date:
            return filter_kline_range(cached, start_date, end_date)
        fetch_start = start_date if cached_max < start_date else (
            pd.to_datetime(cached_max) + pd.DateOffset(days=1)
        ).strftime("%Y-%m-%d")
    else:
        fetch_start = start_date

    if fetch_start <= end_date:
        if source == "akshare":
            fresh = load_kline_akshare(code, fetch_start, end_date, retries=retries, retry_delay=retry_delay)
        else:
            fresh = load_kline_baostock(code, fetch_start, end_date, retries=retries, retry_delay=retry_delay)
        if fresh is not None and not fresh.empty:
            cached = pd.concat([cached, fresh], ignore_index=True, sort=False)
            save_cached_kline(source, code, cached)
    return filter_kline_range(cached, start_date, end_date)


def load_kline_baostock(code, start_date, end_date, retries=4, retry_delay=1.5):
    fields = "date,code,open,high,low,close,volume,tradestatus"
    rs = call_with_backoff(
        lambda: bs.query_history_k_data_plus(
            code,
            fields,
            start_date=start_date,
            end_date=end_date,
            frequency="d",
            adjustflag=ADJUST_FLAG_QFQ,
        ),
        f"{code} baostock K线",
        retries=retries,
        retry_delay=retry_delay,
    )
    df = rs.get_data()
    if rs.error_code != "0" or df.empty:
        return pd.DataFrame()
    df.columns = rs.fields
    if "tradestatus" in df.columns:
        df = df[df["tradestatus"] == "1"]
    return df


def load_kline_akshare(code, start_date, end_date, retries=4, retry_delay=1.5):
    pure = pure_code(code)
    if "." in str(code):
        market = str(code).split(".", 1)[0].lower()
    else:
        market = "sh" if pure.startswith(("6", "9")) else "sz"
    daily_symbol = f"{market}{pure}"

    def fetch_daily():
        return ak.stock_zh_a_daily(
            symbol=daily_symbol,
            start_date=str(start_date).replace("-", ""),
            end_date=str(end_date).replace("-", ""),
            adjust="qfq",
        )

    try:
        df = call_with_backoff(
            fetch_daily,
            f"{code} akshare新浪K线",
            retries=retries,
            retry_delay=retry_delay,
        )
    except Exception as exc:
        print(f"{code} akshare新浪K线失败，回退东方财富: {exc}")
        df = call_with_backoff(
            lambda: ak.stock_zh_a_hist(
                symbol=pure,
                period="daily",
                start_date=str(start_date).replace("-", ""),
                end_date=str(end_date).replace("-", ""),
                adjust="qfq",
            ),
            f"{code} akshare东方财富K线",
            retries=retries,
            retry_delay=retry_delay,
        )
    if df is None or df.empty:
        return pd.DataFrame()
    if "date" in df.columns:
        df = df.copy()
        df["code"] = code
        return df.rename(columns={
            "date": "date",
            "open": "open",
            "high": "high",
            "low": "low",
            "close": "close",
            "volume": "volume",
            "amount": "amount",
        })
    return df.rename(columns={
        "日期": "date",
        "股票代码": "code",
        "开盘": "open",
        "最高": "high",
        "最低": "low",
        "收盘": "close",
        "成交量": "volume",
    })


def fetch_one_stock(task):
    (
        code,
        name,
        start_date,
        end_date,
        date_set,
        mktcap_yi,
        ipo_date,
        min_mktcap,
        min_list_days,
        sleep,
        price_source,
        retries,
        retry_delay,
        debug_filters,
        require_end_trade,
    ) = task
    if sleep > 0:
        time.sleep(sleep)
    if ipo_date is None:
        return []
    if min_mktcap is not None and mktcap_yi is None:
        return []
    try:
        df = load_kline_with_cache(
            price_source,
            code,
            start_date,
            end_date,
            retries=retries,
            retry_delay=retry_delay,
        )
    except Exception as exc:
        return [{"code": code, "name": name, "error": str(exc)}]
    if df.empty:
        return []
    for col in ["open", "high", "low", "close", "volume"]:
        df[col] = pd.to_numeric(df[col], errors="coerce")
    df["date"] = pd.to_datetime(df["date"], errors="coerce").dt.strftime("%Y-%m-%d")
    df = df.dropna(subset=["date", "high", "low", "close"]).sort_values("date").reset_index(drop=True)
    if require_end_trade and df["date"].max() < end_date:
        return []
    if len(df) < 30:
        return []

    k = calc_kdj_k(df)
    wr10 = calc_wr(df, 10)
    wr20 = calc_wr(df, 20)
    rsi9 = calc_rsi(df["close"], 9)
    latest_close = df["close"].iloc[-1]
    if min_mktcap is not None and (not latest_close or pd.isna(latest_close)):
        return []
    ipo_ts = pd.to_datetime(ipo_date, errors="coerce")
    if pd.isna(ipo_ts):
        return []
    current_list_days = (pd.to_datetime(end_date) - ipo_ts).days
    current_mktcap_ok = min_mktcap is None or float(mktcap_yi) > min_mktcap
    current_list_days_ok = current_list_days > min_list_days

    kd80 = k > 80
    wr3 = (wr10 < 20) & (wr20 < 20)
    rsi70 = rsi9 > 70
    base = kd80 & wr3 & rsi70 & current_mktcap_ok & current_list_days_ok
    xg = base.rolling(5, min_periods=5).sum() == 5

    hits = []
    debug_rows = []
    for idx, row in df.iterrows():
        row = df.loc[idx]
        if row["date"] not in date_set:
            continue
        row_date = pd.to_datetime(row["date"], errors="coerce")
        list_days = (row_date - ipo_ts).days if pd.notna(row_date) else current_list_days
        if min_mktcap is None:
            mktcap_at_date = np.nan
        else:
            mktcap_at_date = float(mktcap_yi)
        if debug_filters:
            debug_rows.append({
                "date": row["date"],
                "code": code,
                "name": name,
                "signal_type": "DEBUG",
                "kd80": bool(kd80.loc[idx]),
                "wr3": bool(wr3.loc[idx]),
                "rsi70": bool(rsi70.loc[idx]),
                "mktcap_ok": bool(current_mktcap_ok),
                "list_days_ok": bool(current_list_days_ok),
                "base_ok": bool(base.loc[idx]),
                "xg_ok": bool(xg.loc[idx]),
            })
        if not current_list_days_ok or not current_mktcap_ok:
            continue
        record = {
            "date": row["date"],
            "code": code,
            "name": name,
            "close": row["close"],
            "mktcap_yi": round(mktcap_at_date, 2) if pd.notna(mktcap_at_date) else np.nan,
            "list_days": int(list_days),
            "kdj_k": round(float(k.loc[idx]), 2),
            "wr10": round(float(wr10.loc[idx]), 2),
            "wr20": round(float(wr20.loc[idx]), 2),
            "rsi9": round(float(rsi9.loc[idx]), 2),
        }
        if bool(base.loc[idx]):
            base_record = record.copy()
            base_record["signal_type"] = "BASE"
            hits.append(base_record)
        if bool(xg.loc[idx]):
            xg_record = record.copy()
            xg_record["signal_type"] = "XG"
            hits.append(xg_record)
    return hits + debug_rows


def calc_streaks(counts):
    rows = []
    up_streak = 0
    down_streak = 0
    prev = None
    for date, count in counts:
        change = 0 if prev is None else count - prev
        if prev is None or change == 0:
            up_streak = 0
            down_streak = 0
        elif change > 0:
            up_streak += 1
            down_streak = 0
        else:
            down_streak += 1
            up_streak = 0
        if up_streak >= 5:
            signal = "结构转好确认，右侧成功率提升"
        elif up_streak >= 3:
            signal = "结构初步转好"
        elif down_streak >= 5:
            signal = "结构转坏确认，右侧成功率下降"
        elif down_streak >= 3:
            signal = "结构初步转坏"
        else:
            signal = "观察"
        rows.append({
            "date": date,
            "count": count,
            "change": change,
            "up_streak": up_streak,
            "down_streak": down_streak,
            "signal": signal,
        })
        prev = count
    return pd.DataFrame(rows)


def save_workbook(summary, hits, sample=False, unique_hits=None):
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    suffix = "_sample" if sample else ""
    path = os.path.join(OUTPUT_DIR, f"formula33_stats_{stamp}{suffix}.xlsx")
    csv_path = os.path.join(OUTPUT_DIR, f"formula33_stats_{stamp}{suffix}.csv")
    summary.to_csv(csv_path, index=False, encoding="utf-8-sig")

    wb = Workbook()
    ws = wb.active
    ws.title = "33公式日统计"
    headers = ["date", "base_count", "count", "change", "up_streak", "down_streak", "signal"]
    ws.append(headers)
    for row in summary[headers].to_dict("records"):
        ws.append([row[h] for h in headers])

    ws2 = wb.create_sheet("横向统计")
    ws2.append(["指标"] + summary["date"].tolist())
    ws2.append(["BASE数量"] + summary.get("base_count", pd.Series([np.nan] * len(summary))).tolist())
    ws2.append(["XG数量"] + summary["count"].tolist())
    ws2.append(["较前日变化"] + summary["change"].tolist())
    ws2.append(["连续上升"] + summary["up_streak"].tolist())
    ws2.append(["连续下降"] + summary["down_streak"].tolist())
    ws2.append(["结构信号"] + summary["signal"].tolist())

    ws3 = wb.create_sheet("命中股票")
    if hits.empty:
        ws3.append(["signal_type", "date", "code", "name", "close", "mktcap_yi", "list_days", "kdj_k", "wr10", "wr20", "rsi9"])
    else:
        cols = ["signal_type", "date", "code", "name", "close", "mktcap_yi", "list_days", "kdj_k", "wr10", "wr20", "rsi9"]
        ws3.append(cols)
        for row in hits[cols].to_dict("records"):
            ws3.append([row.get(col) for col in cols])

    ws4 = wb.create_sheet("最近XG去重")
    unique_cols = ["date", "code", "name", "close", "mktcap_yi", "list_days", "kdj_k", "wr10", "wr20", "rsi9"]
    ws4.append(unique_cols)
    if unique_hits is not None and not unique_hits.empty:
        for row in unique_hits[unique_cols].to_dict("records"):
            ws4.append([row.get(col) for col in unique_cols])

    yellow = PatternFill("solid", fgColor="FFF2CC")
    green = PatternFill("solid", fgColor="E2F0D9")
    red = PatternFill("solid", fgColor="F4CCCC")
    thin = Side(style="thin", color="666666")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for sheet in wb.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                if cell.row == 1:
                    cell.font = Font(bold=True)
                    cell.fill = yellow
                if isinstance(cell.value, str) and "转好" in cell.value:
                    cell.fill = green
                if isinstance(cell.value, str) and "转坏" in cell.value:
                    cell.fill = red
        for col in range(1, sheet.max_column + 1):
            sheet.column_dimensions[get_column_letter(col)].width = 18
    wb.save(path)
    return path, csv_path


def build_sample(lookback):
    dates = pd.bdate_range(end=pd.Timestamp("2026-06-24"), periods=lookback).strftime("%Y-%m-%d").tolist()
    counts = [42, 45, 48, 52, 57, 60, 58, 55, 51, 48, 46, 49, 53, 58, 64, 70, 74, 72, 69, 66, 63][-lookback:]
    summary = calc_streaks(list(zip(dates, counts)))
    summary["base_count"] = [value + 80 for value in counts]
    summary = summary[["date", "base_count", "count", "change", "up_streak", "down_streak", "signal"]]
    hit_rows = []
    for date, count in zip(dates, counts):
        for idx in range(min(count, 12)):
            hit_rows.append({
                "signal_type": "XG",
                "date": date,
                "code": f"sz.30{idx:04d}",
                "name": f"33样本{idx + 1}",
                "close": 10 + idx,
                "mktcap_yi": 120 + idx * 5,
                "list_days": 500 + idx,
                "kdj_k": 82 + idx % 10,
                "wr10": 8 + idx % 8,
                "wr20": 9 + idx % 7,
                "rsi9": 72 + idx % 12,
            })
    return summary, pd.DataFrame(hit_rows)


def main():
    args = parse_args()
    if args.sample:
        summary, hits = build_sample(args.lookback)
        xlsx_path, csv_path = save_workbook(summary, hits, sample=True)
        print(f"Excel已保存: {xlsx_path}")
        print(f"CSV已保存: {csv_path}")
        print(summary.tail(8).to_string(index=False))
        return

    lg = bs.login()
    bs_available = lg.error_code == "0"
    if not bs_available:
        print(f"Baostock不可用，改用 akshare 元数据: {lg.error_msg}")
        if args.price_source == "baostock":
            raise SystemExit("Baostock不可用时不能使用 --price-source baostock，请改用 --price-source akshare")
    try:
        raw_trade_dates = get_trade_dates(args.lookback + 5, args.history_days + 45) if bs_available else []
        if len(raw_trade_dates) < args.lookback:
            raw_trade_dates = get_trade_dates_akshare(args.lookback + 5, args.history_days + 45)
        if not raw_trade_dates or (not args.start_date and len(raw_trade_dates) < args.lookback):
            raise SystemExit("交易日不足，无法统计")
        trade_dates = select_trade_dates(raw_trade_dates, args.start_date, args.end_date, args.lookback)
        if not trade_dates or (not args.start_date and len(trade_dates) < args.lookback):
            raise SystemExit(f"交易日不足，无法统计最近 {args.lookback} 个交易日")
        latest_date = trade_dates[-1]
        if bs_available:
            universe_date, universe = get_universe_with_fallback(trade_dates)
        else:
            universe_date, universe = latest_date, get_universe_akshare()
        if universe.empty:
            raise SystemExit("无法获取沪深A股股票池")
        if universe_date != latest_date:
            print(f"股票池使用 {universe_date}，统计交易日仍使用 {latest_date} 之前最近 {len(trade_dates)} 个交易日")
            trade_dates = select_trade_dates(raw_trade_dates, args.start_date, universe_date, args.lookback)
            if not trade_dates or (not args.start_date and len(trade_dates) < args.lookback):
                raise SystemExit(f"股票池最新日期为 {universe_date}，但交易日不足")
            latest_date = trade_dates[-1]
            print(f"已按股票池最新日期裁剪统计窗口，最新统计日: {latest_date}")
        print(f"本次统计区间: {trade_dates[0]} ~ {trade_dates[-1]}，共 {len(trade_dates)} 个交易日")
        start_date = (pd.to_datetime(trade_dates[0]) - pd.DateOffset(days=args.history_days)).strftime("%Y-%m-%d")
        exclude_codes = parse_code_set(args.exclude_codes)
        if exclude_codes:
            before_exclude = len(universe)
            universe = universe[~universe["code"].isin(exclude_codes)].reset_index(drop=True)
            print(f"已排除股票 {len(exclude_codes)} 只，股票池 {before_exclude} -> {len(universe)}")
        if args.offset:
            universe = universe.iloc[args.offset:].reset_index(drop=True)
        if args.limit > 0:
            universe = universe.head(args.limit)
        basic = load_stock_basic() if bs_available else pd.DataFrame()
        if basic.empty:
            basic = load_stock_basic_akshare()
        list_date_map = {}
        if not basic.empty and "code" in basic.columns and "ipoDate" in basic.columns:
            list_date_map = dict(zip(basic["code"], basic["ipoDate"]))
        try:
            cap_map, cap_source = load_market_caps(
                args.market_cap_source,
                latest_date,
                universe,
                capital_workers=args.capital_workers,
                capital_sleep=args.capital_sleep,
                retries=args.retries,
                retry_delay=args.retry_delay,
            )
        except Exception as exc:
            raise SystemExit(f"{exc}\n该公式需要 FINANCE(40)>100亿；可配置 Tushare token，或网络可用时使用 akshare。")
        min_mktcap = None if cap_source == "none" else args.min_mktcap

        date_set = set(trade_dates)
        tasks = []
        for _, row in universe.iterrows():
            code = row["code"]
            ipo = list_date_map.get(code)
            tasks.append((
                code,
                row.get("code_name", ""),
                start_date,
                latest_date,
                date_set,
                cap_map.get(code),
                ipo,
                min_mktcap,
                args.min_list_days,
                args.sleep,
                args.price_source,
                args.retries,
                args.retry_delay,
                args.debug_filters,
                args.require_end_trade,
            ))

        hits = []
        workers = max(1, args.workers)
        print(
            f"候选股票: {len(tasks)} | offset={args.offset} | limit={args.limit} | "
            f"workers={workers} | price_source={args.price_source}"
        )
        if workers > 1:
            if bs_available:
                bs.logout()
            initializer = init_worker if args.price_source == "baostock" else None
            with multiprocessing.Pool(
                processes=workers,
                initializer=initializer,
                maxtasksperchild=args.maxtasksperchild if args.maxtasksperchild > 0 else None,
            ) as pool:
                for idx, result in enumerate(pool.imap_unordered(fetch_one_stock, tasks), start=1):
                    hits.extend([item for item in result if not item.get("error")])
                    if idx % 200 == 0:
                        print(f"进度 {idx}/{len(tasks)}，命中记录 {len(hits)}")
        else:
            for idx, task in enumerate(tasks, start=1):
                hits.extend([item for item in fetch_one_stock(task) if not item.get("error")])
                if idx % 200 == 0:
                    print(f"进度 {idx}/{len(tasks)}，命中记录 {len(hits)}")
            if bs_available:
                bs.logout()

        hits_df = pd.DataFrame(hits)
        if args.debug_filters and not hits_df.empty and "signal_type" in hits_df.columns:
            debug_df = hits_df[hits_df["signal_type"] == "DEBUG"].copy()
            if not debug_df.empty:
                latest_debug = debug_df[debug_df["date"] == latest_date]
                print("\n--- 今日33公式分步诊断 ---")
                for col in ["kd80", "wr3", "rsi70", "mktcap_ok", "list_days_ok", "base_ok", "xg_ok"]:
                    if col in latest_debug.columns:
                        print(f"{col}: {int(latest_debug[col].fillna(False).sum())}")
                hits_df = hits_df[hits_df["signal_type"] != "DEBUG"].copy()
        if hits_df.empty:
            counts = [(date, 0) for date in trade_dates]
            base_counts_by_date = {}
            window_base_unique = 0
            window_xg_unique = 0
            unique_xg_hits = pd.DataFrame()
        else:
            xg_hits = hits_df[hits_df["signal_type"] == "XG"]
            base_hits = hits_df[hits_df["signal_type"] == "BASE"]
            counts_by_date = xg_hits.groupby("date").size().to_dict()
            base_counts_by_date = base_hits.groupby("date").size().to_dict()
            counts = [(date, int(counts_by_date.get(date, 0))) for date in trade_dates]
            window_base_unique = int(base_hits[base_hits["date"].isin(trade_dates)]["code"].nunique())
            window_xg_hits = xg_hits[xg_hits["date"].isin(trade_dates)].copy()
            unique_xg_hits = (
                window_xg_hits.sort_values(["code", "date"])
                .drop_duplicates("code", keep="last")
                .sort_values("code")
            )
            window_xg_unique = int(unique_xg_hits["code"].nunique())
        summary = calc_streaks(counts)
        summary["base_count"] = summary["date"].map(lambda d: int(base_counts_by_date.get(d, 0)))
        summary = summary[["date", "base_count", "count", "change", "up_streak", "down_streak", "signal"]]
        xlsx_path, csv_path = save_workbook(summary, hits_df, unique_hits=unique_xg_hits)
        print(f"Excel已保存: {xlsx_path}")
        print(f"CSV已保存: {csv_path}")
        print(summary.to_string(index=False))
        print(f"最近{len(trade_dates)}个交易日BASE去重股票数: {window_base_unique}")
        print(f"最近{len(trade_dates)}个交易日XG去重股票数: {window_xg_unique}")
    finally:
        if bs_available:
            try:
                bs.logout()
            except Exception:
                pass


if __name__ == "__main__":
    main()
