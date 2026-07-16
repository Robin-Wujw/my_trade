# -*- coding: utf-8 -*-
"""
33 formula market-structure statistics.

Formula:
KD := (CLOSE - LLV(LOW, 9)) / (HHV(HIGH, 9) - LLV(LOW, 9)) * 100;
K := SMA(KD, 3, 1);
WR1 := 100 * (HHV(HIGH, 10) - CLOSE) / (HHV(HIGH, 10) - LLV(LOW, 10));
WR2 := 100 * (HHV(HIGH, 20) - CLOSE) / (HHV(HIGH, 20) - LLV(LOW, 20));
KD80 := K > 80;
WR3 := WR1 < 20 AND WR2 < 20;
RSI70 := SMA(MAX(CLOSE - REF(CLOSE, 1), 0), 9, 1)
    / SMA(ABS(CLOSE - REF(CLOSE, 1)), 9, 1) * 100 > 70;
LIST_DAYS := FINANCE(42) > 300;
BASE := KD80 AND WR3 AND RSI70 AND LIST_DAYS;
XG: COUNT(BASE, 5) = 5;

Market capitalization above CNY 10 billion is exported as a separate
reference pool. It does not filter the formal technical-XG result.

The script records the number of Shanghai/Shenzhen A shares matching XG on
the latest N trading days. Rising for 3/5 days means initial/confirmed
structure improvement; falling for 3/5 days means initial/confirmed weakness.
"""
import argparse
from concurrent.futures import ThreadPoolExecutor, as_completed
import json
import multiprocessing
import os
import time
from datetime import datetime, timedelta

from stock_research.api import akshare as ak
from stock_research.api import baostock as bs
from stock_research.api import tushare as ts_api
from stock_research.api.retry import call_with_backoff
from stock_research.api.schema import rename_columns_strict
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from stock_research.core.completion_manifest import CompletionManifest
from stock_research.core.paths import PATHS
from stock_research.indicators.formula33 import calc_kdj_k, calc_rsi, calc_wr
from stock_research.storage import Database, KlineRepository
from stock_research.strategies.formula33 import (
    build_window_trend,
    classify_observation_status,
    select_window_unique_hits,
)


if not hasattr(pd.DataFrame, "append"):
    def _df_append(self, other, ignore_index=False, **kwargs):
        return pd.concat([self, other], ignore_index=ignore_index)

    pd.DataFrame.append = _df_append


OUTPUT_DIR = str(PATHS.market_exports)
ADJUST_FLAG_QFQ = "2"
SHARE_CACHE_FILE = str(PATHS.cache / "formula33_share_capital.json")
KLINE_CACHE_DIR = str(PATHS.cache / "formula33_kline")
OBSERVATION_SPOT_CACHE_DIR = str(PATHS.cache / "formula33_observation_spot")
FORMULA33_SNAPSHOT_DIR = str(PATHS.cache / "formula33_snapshots")
TRADE_CALENDAR_CACHE_FILE = os.path.join(
    FORMULA33_SNAPSHOT_DIR, "trade_calendar.json"
)
UNIVERSE_CACHE_FILE = str(PATHS.cache / "stock_universe.csv")
FORMULA33_MANIFEST_FILE = str(PATHS.state / "formula33_completion.json")
FORMULA33_CODE_VERSION = "formula33-v7"
KLINE_QFQ_CACHE_VERSION = "qfq-cache-v2"
REQUEST_RETRY_ERRORS = (BrokenPipeError, ConnectionError, TimeoutError, OSError)
MIN_LIST_DATE_COVERAGE = 0.98
MIN_MARKET_CAP_COVERAGE = 0.98
MIN_OBSERVATION_STATUS_COVERAGE = 0.98


def parse_args(argv=None):
    parser = argparse.ArgumentParser(description="33公式沪深A股市场结构统计")
    parser.add_argument("--lookback", type=int, default=21, help="统计最近N个交易日")
    parser.add_argument("--start-date", default="", help="统计起始交易日，格式 YYYY-MM-DD；传入后按日期区间统计")
    parser.add_argument("--end-date", default="", help="统计截止交易日，格式 YYYY-MM-DD；留空则使用当前日期")
    parser.add_argument("--data-ready-time", default="16:00", help="未指定 --end-date 时，当天日线视为可用的本地时间")
    parser.add_argument("--history-days", type=int, default=90, help="为指标计算额外拉取的自然日长度")
    parser.add_argument("--workers", type=int, default=4)
    parser.add_argument("--sleep", type=float, default=0.0)
    parser.add_argument("--retries", type=int, default=4, help="单只股票行情请求失败重试次数")
    parser.add_argument("--retry-delay", type=float, default=1.5, help="请求失败后的退避基准秒数")
    parser.add_argument("--capital-workers", type=int, default=2, help="沪市股本结构补充并发数")
    parser.add_argument("--capital-sleep", type=float, default=0.08, help="沪市股本结构补充每个请求前等待秒数")
    parser.add_argument("--offset", type=int, default=0, help="分段续跑用，从股票池第N只开始")
    parser.add_argument("--limit", type=int, default=0, help="调试用，只处理前N只")
    parser.add_argument("--exclude-codes", default="", help="逗号分隔的股票代码，按截图/外部股票池复核时可排除")
    parser.add_argument("--maxtasksperchild", type=int, default=200, help="多进程模式下每个worker处理多少任务后重启")
    parser.add_argument(
        "--price-source",
        choices=["tushare", "akshare", "baostock"],
        default="akshare",
        help="前复权K线来源；默认 AkShare，Tushare 低频账号仅适合显式小批量使用",
    )
    parser.add_argument("--metadata-source", choices=["akshare", "baostock", "auto"], default="akshare", help="交易日/股票池/上市日期来源，默认优先 AkShare")
    parser.add_argument("--min-mktcap", type=float, default=100.0, help="最低总市值，单位亿元")
    parser.add_argument("--min-list-days", type=int, default=300, help="最低上市天数")
    parser.add_argument("--debug-filters", action="store_true", help="打印最近交易日各条件逐步通过数量")
    parser.add_argument("--require-end-trade", action="store_true", help="只保留截止日有K线的股票，用于复核当前交易列表并排除停牌票")
    parser.add_argument(
        "--market-cap-source",
        choices=["auto", "tushare", "akshare", "akshare-capital", "none"],
        default="auto",
        help="总市值来源；none 仅用于临时复核技术指标数量，会跳过 FINANCE(40) 过滤",
    )
    parser.add_argument(
        "--missing-mktcap-policy",
        choices=["exclude", "pass"],
        default="pass",
        help="总市值源缺失个股时的处理：exclude 严格剔除，pass 不因接口缺字段误杀",
    )
    parser.add_argument("--sample", action="store_true", help="生成离线样例，不访问网络")
    return parser.parse_args(argv)


def _atomic_write_csv(frame, path):
    os.makedirs(os.path.dirname(os.path.abspath(path)), exist_ok=True)
    temporary = f"{path}.{os.getpid()}.{time.time_ns()}.tmp"
    try:
        frame.to_csv(temporary, index=False, encoding="utf-8-sig")
        os.replace(temporary, path)
    except Exception:
        try:
            os.remove(temporary)
        except FileNotFoundError:
            pass
        raise


def _atomic_write_json(payload, path):
    os.makedirs(os.path.dirname(os.path.abspath(path)), exist_ok=True)
    temporary = f"{path}.{os.getpid()}.{time.time_ns()}.tmp"
    try:
        with open(temporary, "w", encoding="utf-8") as handle:
            json.dump(payload, handle, ensure_ascii=False, sort_keys=True, indent=2)
            handle.write("\n")
            handle.flush()
            os.fsync(handle.fileno())
        os.replace(temporary, path)
    except Exception:
        try:
            os.remove(temporary)
        except FileNotFoundError:
            pass
        raise


def _read_json(path):
    try:
        with open(path, "r", encoding="utf-8") as handle:
            payload = json.load(handle)
    except (OSError, ValueError, TypeError):
        return {}
    return payload if isinstance(payload, dict) else {}


def _snapshot_path(kind, observation_date, suffix):
    date_key = pd.Timestamp(observation_date).strftime("%Y%m%d")
    return os.path.join(FORMULA33_SNAPSHOT_DIR, f"{kind}_{date_key}.{suffix}")


def _lookup_coverage(universe, lookup):
    universe_codes = set(universe["code"].dropna().astype(str))
    covered = {
        str(code)
        for code, value in dict(lookup or {}).items()
        if str(code) in universe_codes
        and value is not None
        and not pd.isna(value)
        and str(value).strip()
    }
    return len(covered) / len(universe_codes) if universe_codes else 0.0


def to_bs_code(raw_code):
    code = str(raw_code).strip()
    if "." in code:
        return code
    low = code.lower()
    if low.startswith("sh") and len(code) >= 8:
        return f"sh.{code[-6:]}"
    if low.startswith("sz") and len(code) >= 8:
        return f"sz.{code[-6:]}"
    if low.startswith("bj") and len(code) >= 8:
        return f"bj.{code[-6:]}"
    code = code.zfill(6)
    if code.startswith(("6", "9")):
        return f"sh.{code}"
    return f"sz.{code}"


def pure_code(bs_code):
    return str(bs_code).split(".")[-1]


def normalize_six_digit_code(raw_code):
    text = str(raw_code).strip()
    if text.endswith(".0"):
        text = text[:-2]
    digits = "".join(ch for ch in text if ch.isdigit())
    return digits[-6:].zfill(6) if digits else ""


def get_trade_dates(lookback, extra_days):
    end = pd.Timestamp.today().strftime("%Y-%m-%d")
    start = (pd.Timestamp.today() - pd.DateOffset(days=extra_days)).strftime("%Y-%m-%d")
    rs = bs.query_trade_dates(start_date=start, end_date=end)
    df = rs.get_data()
    if rs.error_code != "0" or df.empty:
        return []
    df.columns = rs.fields
    df = df[df["is_trading_day"] == "1"]
    return df["calendar_date"].tail(lookback).tolist()


def _calendar_snapshot_is_reusable(payload, required_through):
    dates = payload.get("trade_dates")
    if not isinstance(dates, list) or not dates:
        return False
    parsed_dates = pd.to_datetime(pd.Series(dates), errors="coerce").dropna()
    if parsed_dates.empty:
        return False
    required = pd.Timestamp(required_through).normalize()
    if required <= parsed_dates.max().normalize():
        return True
    fetched_for = pd.to_datetime(payload.get("fetched_for_date"), errors="coerce")
    if pd.isna(fetched_for):
        return False
    fetched_for = fetched_for.normalize()
    if fetched_for == required:
        return True
    day_gap = (required - fetched_for).days
    return (
        fetched_for.weekday() == 4
        and required.weekday() in (5, 6)
        and day_gap == required.weekday() - fetched_for.weekday()
    )


def get_trade_dates_akshare(lookback, extra_days, required_through=None):
    required = pd.Timestamp(required_through or pd.Timestamp.today()).normalize()
    payload = _read_json(TRADE_CALENDAR_CACHE_FILE)
    if _calendar_snapshot_is_reusable(payload, required):
        all_dates = payload["trade_dates"]
        cache_status = "hit"
    else:
        df = ak.tool_trade_date_hist_sina()
        if df is None or df.empty or "trade_date" not in df.columns:
            return []
        parsed = pd.to_datetime(df["trade_date"], errors="coerce").dropna()
        all_dates = sorted(set(parsed.dt.strftime("%Y-%m-%d")))
        if not all_dates:
            return []
        payload = {
            "version": 1,
            "fetched_for_date": required.strftime("%Y-%m-%d"),
            "trade_dates": all_dates,
        }
        _atomic_write_json(payload, TRADE_CALENDAR_CACHE_FILE)
        cache_status = "write"
    end = required
    start = end - pd.DateOffset(days=extra_days)
    selected = [
        value
        for value in all_dates
        if start <= pd.Timestamp(value).normalize() <= end
    ]
    print(
        f"Formula33 trade-calendar snapshot {cache_status}: "
        f"through={required.strftime('%Y-%m-%d')} rows={len(selected)}"
    )
    return selected[-lookback:]


def latest_trade_date_from_calendar_snapshot(as_of_date=None):
    payload = _read_json(TRADE_CALENDAR_CACHE_FILE)
    dates = pd.to_datetime(
        pd.Series(payload.get("trade_dates", []), dtype="object"),
        errors="coerce",
    ).dropna()
    if dates.empty:
        return ""
    as_of = pd.Timestamp(as_of_date or pd.Timestamp.today()).normalize()
    dates = dates[dates.dt.normalize() <= as_of]
    if dates.empty:
        return ""
    return dates.max().strftime("%Y-%m-%d")


def select_trade_dates(trade_dates, start_date, end_date, lookback):
    if start_date:
        start = pd.to_datetime(start_date, errors="coerce")
        if pd.isna(start):
            raise SystemExit(f"--start-date 格式错误: {start_date}")
        start_text = start.strftime("%Y-%m-%d")
        trade_dates = [date for date in trade_dates if date >= start_text]
    if end_date:
        cap = pd.to_datetime(end_date, errors="coerce")
        if pd.isna(cap):
            raise SystemExit(f"--end-date 格式错误: {end_date}")
        cap_text = cap.strftime("%Y-%m-%d")
        trade_dates = [date for date in trade_dates if date <= cap_text]
    if start_date:
        return trade_dates
    return trade_dates[-lookback:]


def resolve_auto_end_date(end_date, data_ready_time):
    if end_date:
        return end_date
    try:
        hour, minute = [int(part) for part in str(data_ready_time).split(":", 1)]
    except (TypeError, ValueError):
        raise SystemExit(f"--data-ready-time 格式错误: {data_ready_time}")
    now = datetime.now()
    ready_at = now.replace(hour=hour, minute=minute, second=0, microsecond=0)
    if now < ready_at:
        return (now.date() - timedelta(days=1)).strftime("%Y-%m-%d")
    return now.strftime("%Y-%m-%d")


def parse_code_set(value):
    codes = set()
    for item in str(value or "").replace("，", ",").split(","):
        item = item.strip()
        if item:
            codes.add(to_bs_code(item))
    return codes


def build_completion_arguments(args, observation_date):
    """Return only result-affecting arguments in their effective form."""
    normalized_start_date = (
        pd.Timestamp(args.start_date).strftime("%Y-%m-%d")
        if args.start_date
        else ""
    )
    return {
        "lookback": int(args.lookback),
        "start_date": normalized_start_date,
        "end_date": str(observation_date),
        "history_days": int(args.history_days),
        "offset": int(args.offset),
        "limit": int(args.limit),
        "exclude_codes": sorted(parse_code_set(args.exclude_codes)),
        "price_source": str(args.price_source),
        "metadata_source": str(args.metadata_source),
        "min_mktcap": float(args.min_mktcap),
        "min_list_days": int(args.min_list_days),
        "require_end_trade": bool(args.require_end_trade),
        "market_cap_source": str(args.market_cap_source),
        "missing_mktcap_policy": str(args.missing_mktcap_policy),
    }


def load_cached_universe():
    try:
        cached = pd.read_csv(UNIVERSE_CACHE_FILE, dtype=str)
    except (OSError, ValueError, pd.errors.ParserError):
        return pd.DataFrame()
    if "code" not in cached.columns:
        return pd.DataFrame()
    cached = cached.dropna(subset=["code"]).copy()
    cached["code"] = cached["code"].astype(str).str.strip()
    return cached[cached["code"] != ""].drop_duplicates("code").reset_index(drop=True)


def select_universe_for_run(universe, args, announce=False):
    selected = universe.copy()
    exclude_codes = parse_code_set(args.exclude_codes)
    if exclude_codes:
        before_exclude = len(selected)
        selected = selected[~selected["code"].isin(exclude_codes)].reset_index(drop=True)
        if announce:
            print(
                f"已排除股票 {len(exclude_codes)} 只，股票池 "
                f"{before_exclude} -> {len(selected)}"
            )
    if args.offset:
        selected = selected.iloc[args.offset:].reset_index(drop=True)
    if args.limit > 0:
        selected = selected.head(args.limit).reset_index(drop=True)
    return selected


def reuse_completed_manifest_without_network(args):
    manifest = CompletionManifest(FORMULA33_MANIFEST_FILE)
    payload = manifest.read()
    observation_text = str(payload.get("observation_date", "")).strip()
    observation = pd.to_datetime(observation_text, errors="coerce")
    if pd.isna(observation):
        return False
    requested = pd.to_datetime(
        resolve_auto_end_date(args.end_date, args.data_ready_time),
        errors="coerce",
    )
    if pd.isna(requested):
        return False
    requested = requested.normalize()
    observation = observation.normalize()
    day_gap = (requested - observation).days
    locally_safe = day_gap == 0 or (
        observation.weekday() == 4
        and requested.weekday() in (5, 6)
        and day_gap == requested.weekday() - observation.weekday()
    )
    if not locally_safe:
        return False
    cached_universe = load_cached_universe()
    if cached_universe.empty:
        return False
    cached_universe = select_universe_for_run(cached_universe, args)
    if not manifest.matches(
        observation_date=observation_text,
        arguments=build_completion_arguments(args, observation_text),
        universe_codes=cached_universe["code"].astype(str).tolist(),
        code_version=FORMULA33_CODE_VERSION,
    ):
        return False
    print(
        f"Formula33 resume: completed manifest hit date={observation_text}; "
        "network_fetch=0"
    )
    return True


def require_lookup_coverage(label, universe, lookup, *, minimum):
    universe_codes = set(universe["code"].dropna().astype(str))
    available_codes = {
        str(code)
        for code, value in dict(lookup or {}).items()
        if str(code) in universe_codes
        and value is not None
        and not pd.isna(value)
        and str(value).strip()
    }
    coverage = len(available_codes) / len(universe_codes) if universe_codes else 0.0
    if not universe_codes or coverage < float(minimum):
        raise RuntimeError(
            f"{label}覆盖率不足: covered={len(available_codes)} "
            f"universe={len(universe_codes)} coverage={coverage:.1%} "
            f"minimum={float(minimum):.1%}"
        )
    return coverage


def observation_spot_cache_path(observation_date):
    date_key = pd.Timestamp(observation_date).strftime("%Y%m%d")
    return os.path.join(OBSERVATION_SPOT_CACHE_DIR, f"spot_{date_key}.csv")


def _valid_observation_spot(snapshot, observation_date):
    required = {"代码", "今开", "最高", "最低", "成交量", "成交额"}
    if snapshot is None or snapshot.empty or not required.issubset(snapshot.columns):
        return False
    if "observation_date" not in snapshot.columns:
        return False
    dates = set(snapshot["observation_date"].dropna().astype(str))
    return dates == {str(observation_date)}


def load_observation_spot_snapshot(observation_date, *, allow_network):
    path = observation_spot_cache_path(observation_date)
    try:
        cached = pd.read_csv(path, dtype={"代码": str})
    except (FileNotFoundError, OSError, ValueError, pd.errors.ParserError):
        cached = pd.DataFrame()
    if _valid_observation_spot(cached, observation_date):
        return cached
    if not allow_network:
        return pd.DataFrame()

    snapshot = ak.stock_zh_a_spot()
    if snapshot is None or snapshot.empty:
        raise RuntimeError("akshare stock_zh_a_spot returned no observation snapshot")
    snapshot = snapshot.copy()
    snapshot["observation_date"] = str(observation_date)
    if not _valid_observation_spot(snapshot, observation_date):
        raise RuntimeError("observation snapshot is missing required quote fields")

    os.makedirs(os.path.dirname(path), exist_ok=True)
    tmp_path = f"{path}.{os.getpid()}.tmp"
    try:
        snapshot.to_csv(tmp_path, index=False, encoding="utf-8-sig")
        os.replace(tmp_path, path)
    except Exception as exc:
        try:
            if os.path.exists(tmp_path):
                os.remove(tmp_path)
        except OSError:
            pass
        raise RuntimeError(f"observation snapshot cache write failed: {exc}") from exc
    return snapshot


def build_observation_trade_status(universe, snapshot, *, minimum=0.98):
    universe_codes = set(universe["code"].dropna().astype(str))
    statuses = {code: "unknown" for code in universe_codes}
    if snapshot is None or snapshot.empty:
        return statuses, 0.0

    data = snapshot.copy()
    data["code"] = data["代码"].map(to_bs_code)
    numeric_columns = ["今开", "最高", "最低", "成交量", "成交额"]
    for column in numeric_columns:
        data[column] = pd.to_numeric(data[column], errors="coerce")
    data = data[data["code"].isin(universe_codes)].drop_duplicates("code", keep="last")
    conclusive = 0
    zero_columns = ["今开", "最高", "最低", "成交量", "成交额"]
    for _, row in data.iterrows():
        code = row["code"]
        values = [row[column] for column in zero_columns]
        if all(pd.notna(value) and float(value) == 0.0 for value in values):
            statuses[code] = "suspended"
            conclusive += 1
        elif (
            pd.notna(row["今开"])
            and pd.notna(row["最高"])
            and pd.notna(row["最低"])
            and pd.notna(row["成交量"])
            and pd.notna(row["成交额"])
            and float(row["今开"]) > 0
            and float(row["最高"]) > 0
            and float(row["最低"]) > 0
            and float(row["成交量"]) > 0
            and float(row["成交额"]) > 0
        ):
            statuses[code] = "traded"
            conclusive += 1

    coverage = conclusive / len(universe_codes) if universe_codes else 0.0
    if not universe_codes or coverage < float(minimum):
        raise RuntimeError(
            "observation trade-status coverage insufficient: "
            f"covered={conclusive} universe={len(universe_codes)} "
            f"coverage={coverage:.1%} minimum={float(minimum):.1%}"
        )
    return statuses, coverage


def get_universe(latest_date):
    rs = bs.query_all_stock(day=latest_date)
    df = rs.get_data()
    if rs.error_code != "0" or df.empty:
        return pd.DataFrame()
    df.columns = rs.fields
    mask = (
        df["code"].str.startswith("sh.60")
        | df["code"].str.startswith("sh.68")
        | df["code"].str.startswith("sz.00")
        | df["code"].str.startswith("sz.30")
    )
    df = df[mask].copy()
    if "tradeStatus" in df.columns:
        df = df[df["tradeStatus"] != "0"]
    if "code_name" in df.columns:
        df = df[~df["code_name"].astype(str).str.contains("ST", na=False)]
    result = df[["code", "code_name"]].drop_duplicates("code").reset_index(drop=True)
    try:
        os.makedirs(os.path.dirname(UNIVERSE_CACHE_FILE), exist_ok=True)
        result.to_csv(UNIVERSE_CACHE_FILE, index=False, encoding="utf-8-sig")
    except OSError:
        pass
    return result


def get_universe_with_fallback(trade_dates):
    for date in reversed(trade_dates):
        df = get_universe(date)
        if not df.empty:
            return date, df
    return "", pd.DataFrame()


def get_universe_akshare():
    df = ak.stock_info_a_code_name()
    if df is None or df.empty:
        return pd.DataFrame()
    df = rename_columns_strict(
        df,
        {
            "raw_code": ("code", "股票代码", "证券代码", "代码"),
            "code_name": ("name", "股票简称", "证券简称", "名称"),
        },
        label="AkShare A-share universe",
    )
    df["code"] = df["raw_code"].map(to_bs_code)
    mask = (
        df["code"].str.startswith("sh.60")
        | df["code"].str.startswith("sh.68")
        | df["code"].str.startswith("sz.00")
        | df["code"].str.startswith("sz.30")
    )
    df = df[mask].copy()
    df = df[~df["code_name"].astype(str).str.contains("ST|退", na=False)]
    return df[["code", "code_name"]].drop_duplicates("code").reset_index(drop=True)


def _valid_universe_frame(frame):
    return (
        isinstance(frame, pd.DataFrame)
        and not frame.empty
        and {"code", "code_name"}.issubset(frame.columns)
        and frame["code"].notna().all()
    )


def load_universe_snapshot(observation_date):
    path = _snapshot_path("universe", observation_date, "csv")
    try:
        cached = pd.read_csv(path, dtype=str)
    except (OSError, ValueError, pd.errors.ParserError):
        cached = pd.DataFrame()
    if _valid_universe_frame(cached):
        general = load_cached_universe()
        cached_codes = set(cached["code"].astype(str))
        general_codes = (
            set(general["code"].astype(str)) if not general.empty else set()
        )
        if cached_codes != general_codes:
            _atomic_write_csv(cached, UNIVERSE_CACHE_FILE)
        print(
            f"Formula33 universe snapshot hit: date={observation_date} "
            f"rows={len(cached)}"
        )
        return cached.drop_duplicates("code").reset_index(drop=True)

    previous = load_cached_universe()
    fresh = get_universe_akshare()
    if not _valid_universe_frame(fresh):
        return pd.DataFrame()
    if len(previous) >= 100 and len(fresh) < int(np.ceil(len(previous) * 0.98)):
        raise RuntimeError(
            "Formula33 universe snapshot is unexpectedly truncated: "
            f"fresh={len(fresh)} previous={len(previous)}"
        )
    _atomic_write_csv(fresh, path)
    _atomic_write_csv(fresh, UNIVERSE_CACHE_FILE)
    print(
        f"Formula33 universe snapshot write: date={observation_date} "
        f"rows={len(fresh)}"
    )
    return fresh


def load_stock_basic():
    try:
        rs = bs.query_stock_basic()
        df = rs.get_data()
    except Exception:
        return pd.DataFrame()
    if df is None or df.empty:
        return pd.DataFrame()
    df.columns = rs.fields
    return df


def load_stock_basic_akshare():
    rows = []
    errors = []
    for symbol in ["\u4e3b\u677fA\u80a1", "\u79d1\u521b\u677f"]:
        try:
            sh = ak.stock_info_sh_name_code(symbol=symbol)
            sh = rename_columns_strict(
                sh,
                {
                    "raw_code": ("证券代码", "股票代码", "A股代码", "code"),
                    "ipo_date": ("上市日期", "A股上市日期", "list_date", "ipo_date"),
                },
                label=f"AkShare Shanghai stock basic/{symbol}",
            )
            for _, row in sh.iterrows():
                rows.append({
                    "code": f"sh.{str(row.get('raw_code')).zfill(6)}",
                    "ipoDate": row.get("ipo_date"),
                })
        except Exception as exc:
            errors.append(str(exc))
            continue
    try:
        sz = ak.stock_info_sz_name_code(symbol="\u0041\u80a1\u5217\u8868")
        sz = rename_columns_strict(
            sz,
            {
                "raw_code": ("A股代码", "证券代码", "股票代码", "code"),
                "ipo_date": ("A股上市日期", "上市日期", "list_date", "ipo_date"),
            },
            label="AkShare Shenzhen stock basic",
        )
        for _, row in sz.iterrows():
            rows.append({
                "code": f"sz.{str(row.get('raw_code')).zfill(6)}",
                "ipoDate": row.get("ipo_date"),
            })
    except Exception as exc:
        errors.append(str(exc))
    if errors:
        print("AkShare 上市日期部分接口异常: " + " | ".join(errors))
    return pd.DataFrame(rows)


def load_stock_basic_snapshot(observation_date, universe):
    path = _snapshot_path("stock_basic", observation_date, "csv")
    try:
        cached = pd.read_csv(path, dtype={"code": str})
    except (OSError, ValueError, pd.errors.ParserError):
        cached = pd.DataFrame()
    if {"code", "ipoDate"}.issubset(cached.columns):
        cached_lookup = dict(zip(cached["code"], cached["ipoDate"]))
        if _lookup_coverage(universe, cached_lookup) >= MIN_LIST_DATE_COVERAGE:
            print(
                f"Formula33 stock-basic snapshot hit: date={observation_date} "
                f"rows={len(cached)}"
            )
            return cached

    fresh = load_stock_basic_akshare()
    if {"code", "ipoDate"}.issubset(fresh.columns):
        fresh_lookup = dict(zip(fresh["code"], fresh["ipoDate"]))
        if _lookup_coverage(universe, fresh_lookup) >= MIN_LIST_DATE_COVERAGE:
            _atomic_write_csv(fresh, path)
            print(
                f"Formula33 stock-basic snapshot write: date={observation_date} "
                f"rows={len(fresh)}"
            )
    return fresh


def get_tushare_token():
    return ts_api.get_token()


def load_market_caps_from_tushare(trade_date):
    ts_date = str(trade_date).replace("-", "")
    df = ts_api.query(
        "daily_basic",
        trade_date=ts_date,
        fields="ts_code,total_mv",
    )
    if df is None or df.empty:
        raise RuntimeError(f"tushare daily_basic 在 {trade_date} 无数据")
    caps = {}
    for _, row in df.iterrows():
        ts_code = str(row.get("ts_code", ""))
        code = ts_code.split(".")[0]
        if ts_code.endswith(".SH"):
            bs_code = f"sh.{code}"
        elif ts_code.endswith(".SZ"):
            bs_code = f"sz.{code}"
        else:
            bs_code = to_bs_code(code)
        total_mv_wan = pd.to_numeric(row.get("total_mv"), errors="coerce")
        if pd.notna(total_mv_wan):
            caps[bs_code] = float(total_mv_wan) / 10000.0
    return caps


def load_market_caps_from_akshare():
    try:
        df = ak.stock_zh_a_spot_em()
    except Exception as exc:
        raise RuntimeError(f"无法获取东方财富A股总市值：{exc}")
    if df is None or df.empty:
        return {}
    code_col = "代码" if "代码" in df.columns else None
    cap_col = "总市值" if "总市值" in df.columns else None
    if not code_col or not cap_col:
        raise RuntimeError("东方财富行情表缺少 代码/总市值 字段")
    caps = {}
    for _, row in df.iterrows():
        cap = pd.to_numeric(row.get(cap_col), errors="coerce")
        if pd.isna(cap):
            continue
        caps[to_bs_code(row.get(code_col))] = float(cap) / 100000000.0
    return caps


def load_share_cache():
    try:
        if os.path.exists(SHARE_CACHE_FILE):
            with open(SHARE_CACHE_FILE, "r", encoding="utf-8") as fh:
                return json.load(fh)
    except (OSError, json.JSONDecodeError):
        return {}
    return {}


def save_share_cache(cache):
    try:
        os.makedirs(os.path.dirname(SHARE_CACHE_FILE), exist_ok=True)
        with open(SHARE_CACHE_FILE, "w", encoding="utf-8") as fh:
            json.dump(cache, fh, ensure_ascii=False, indent=2)
    except OSError:
        pass


def parse_share(value):
    text = str(value).replace(",", "").strip()
    val = pd.to_numeric(text, errors="coerce")
    return None if pd.isna(val) else float(val)


def em_symbol_from_bs(code):
    pure = pure_code(code)
    if str(code).startswith("sh."):
        return f"{pure}.SH"
    if str(code).startswith("sz."):
        return f"{pure}.SZ"
    return pure


def fetch_total_share_from_gbjg(task):
    code, sleep, retries, retry_delay = task
    if sleep > 0:
        time.sleep(sleep)
    try:
        gb = call_with_backoff(
            lambda: ak.stock_zh_a_gbjg_em(symbol=em_symbol_from_bs(code)),
            f"{code} 股本结构",
            retries=retries,
            retry_delay=retry_delay,
        )
        if gb is None or gb.empty or "总股本" not in gb.columns:
            return code, None
        gb = gb.copy()
        gb["变更日期_dt"] = pd.to_datetime(gb["变更日期"], errors="coerce")
        gb = gb.sort_values("变更日期_dt", ascending=False)
        return code, parse_share(gb.iloc[0].get("总股本"))
    except Exception:
        return code, None


def load_market_caps_from_akshare_capital(
    universe,
    capital_workers=2,
    capital_sleep=0.08,
    retries=4,
    retry_delay=1.5,
    spot_snapshot=None,
):
    spot = (
        spot_snapshot.copy()
        if spot_snapshot is not None and not spot_snapshot.empty
        else ak.stock_zh_a_spot()
    )
    if spot is None or spot.empty:
        raise RuntimeError("akshare stock_zh_a_spot 无数据，无法取得当前价格")
    code_col = "代码"
    price_col = "最新价"
    price_map = {}
    for _, row in spot.iterrows():
        price = pd.to_numeric(row.get(price_col), errors="coerce")
        if pd.notna(price) and price > 0:
            price_map[to_bs_code(row.get(code_col))] = float(price)

    share_map = {}
    try:
        sz = ak.stock_info_sz_name_code(symbol="\u0041\u80a1\u5217\u8868")
        for _, row in sz.iterrows():
            raw_code = normalize_six_digit_code(row.get("A股代码"))
            if not raw_code:
                continue
            code = f"sz.{raw_code}"
            share = parse_share(row.get("A股总股本"))
            if share:
                share_map[code] = share
    except Exception as exc:
        print(f"深市总股本读取失败: {exc}")

    cache = load_share_cache()
    missing_sh = []
    for _, row in universe.iterrows():
        code = row["code"]
        if code in share_map:
            continue
        cached = cache.get(code)
        if cached and cached.get("total_share"):
            share_map[code] = float(cached["total_share"])
            continue
        if not code.startswith("sh."):
            continue
        missing_sh.append(code)
    if missing_sh:
        workers = max(1, capital_workers)
        print(f"沪市总股本需补充 {len(missing_sh)} 只，使用 akshare 股本结构读取，并发 {workers}")
        with ThreadPoolExecutor(max_workers=workers) as executor:
            futures = [
                executor.submit(fetch_total_share_from_gbjg, (code, capital_sleep, retries, retry_delay))
                for code in missing_sh
            ]
            for idx, future in enumerate(as_completed(futures), start=1):
                code, share = future.result()
                if share:
                    share_map[code] = share
                    cache[code] = {"total_share": share, "updated_at": datetime.now().strftime("%Y-%m-%d")}
                if idx % 200 == 0:
                    print(f"  沪市总股本进度 {idx}/{len(missing_sh)}")
    save_share_cache(cache)

    caps = {}
    for code, share in share_map.items():
        price = price_map.get(code)
        if price:
            caps[code] = share * price / 100000000.0
    if not caps:
        raise RuntimeError("akshare 股本结构法未生成有效总市值")
    return caps


def load_market_caps(
    source,
    trade_date,
    universe=None,
    capital_workers=2,
    capital_sleep=0.08,
    retries=4,
    retry_delay=1.5,
    spot_snapshot=None,
):
    if source == "none":
        print("已按 --market-cap-source none 跳过 FINANCE(40) 市值过滤，仅用于技术条件复核。")
        return {}, "none"

    errors = []
    sources = ["tushare", "akshare", "akshare-capital"] if source == "auto" else [source]
    for item in sources:
        try:
            if item == "tushare":
                caps = load_market_caps_from_tushare(trade_date)
            elif item == "akshare":
                caps = load_market_caps_from_akshare()
            elif item == "akshare-capital":
                if universe is None:
                    raise RuntimeError("akshare-capital 需要股票池 universe")
                caps = load_market_caps_from_akshare_capital(
                    universe,
                    capital_workers=capital_workers,
                    capital_sleep=capital_sleep,
                    retries=retries,
                    retry_delay=retry_delay,
                    spot_snapshot=spot_snapshot,
                )
            else:
                continue
            if caps:
                print(f"总市值来源: {item}，记录数 {len(caps)}")
                return caps, item
        except Exception as exc:
            errors.append(f"{item}: {exc}")
    raise RuntimeError("无法获取总市值数据；" + " | ".join(errors))


def load_market_cap_snapshot(
    source,
    trade_date,
    universe,
    *,
    capital_workers=2,
    capital_sleep=0.08,
    retries=4,
    retry_delay=1.5,
    spot_snapshot=None,
):
    if source == "none":
        return load_market_caps(
            source,
            trade_date,
            universe,
            capital_workers=capital_workers,
            capital_sleep=capital_sleep,
            retries=retries,
            retry_delay=retry_delay,
            spot_snapshot=spot_snapshot,
        )
    safe_source = "".join(
        char if char.isalnum() or char in {"-", "_"} else "_"
        for char in str(source)
    )
    path = _snapshot_path(f"market_caps_{safe_source}", trade_date, "json")
    payload = _read_json(path)
    cached_caps = payload.get("market_caps")
    if (
        payload.get("observation_date") == str(trade_date)
        and payload.get("requested_source") == str(source)
        and isinstance(cached_caps, dict)
        and _lookup_coverage(universe, cached_caps) >= MIN_MARKET_CAP_COVERAGE
    ):
        resolved_source = str(payload.get("resolved_source", source))
        caps = {str(code): float(value) for code, value in cached_caps.items()}
        print(
            f"Formula33 market-cap snapshot hit: date={trade_date} "
            f"source={resolved_source} rows={len(caps)}"
        )
        return caps, resolved_source

    caps, resolved_source = load_market_caps(
        source,
        trade_date,
        universe,
        capital_workers=capital_workers,
        capital_sleep=capital_sleep,
        retries=retries,
        retry_delay=retry_delay,
        spot_snapshot=spot_snapshot,
    )
    if _lookup_coverage(universe, caps) >= MIN_MARKET_CAP_COVERAGE:
        _atomic_write_json(
            {
                "version": 1,
                "observation_date": str(trade_date),
                "requested_source": str(source),
                "resolved_source": str(resolved_source),
                "market_caps": {
                    str(code): float(value)
                    for code, value in caps.items()
                    if value is not None and not pd.isna(value)
                },
            },
            path,
        )
        print(
            f"Formula33 market-cap snapshot write: date={trade_date} "
            f"source={resolved_source} rows={len(caps)}"
        )
    return caps, resolved_source


def init_worker():
    bs.ensure_success(bs.login(), "BaoStock worker login")


def kline_cache_path(source, code):
    safe_code = str(code).replace(".", "_")
    return os.path.join(KLINE_CACHE_DIR, source, f"{safe_code}.csv")


def kline_no_trade_marker_path(source, code):
    return f"{kline_cache_path(source, code)}.no-trade.json"


def kline_cache_metadata_path(source, code):
    return f"{kline_cache_path(source, code)}.meta.json"


def _kline_cache_file_signature(source, code):
    try:
        stat = os.stat(kline_cache_path(source, code))
    except OSError:
        return {}
    return {
        "size": int(stat.st_size),
        "mtime_ns": int(stat.st_mtime_ns),
    }


def _kline_anchor_date_from_frame(df):
    normalized = normalize_kline_df(df)
    if normalized.empty:
        return None
    return str(normalized["date"].max())


def _kline_anchor_allowed(payload, max_qfq_anchor_date=None):
    if max_qfq_anchor_date is None:
        return True
    anchor = pd.to_datetime(payload.get("qfq_anchor_date"), errors="coerce")
    maximum = pd.to_datetime(max_qfq_anchor_date, errors="coerce")
    if pd.isna(anchor) or pd.isna(maximum):
        return False
    return anchor.normalize() <= maximum.normalize()


def _legacy_kline_cache_can_infer_anchor(df, max_qfq_anchor_date=None):
    if max_qfq_anchor_date is None:
        return False
    anchor_date = _kline_anchor_date_from_frame(df)
    if not anchor_date:
        return False
    payload = {"qfq_anchor_date": anchor_date}
    return _kline_anchor_allowed(payload, max_qfq_anchor_date)


def _kline_cache_metadata_missing(source, code):
    return not bool(_read_json(kline_cache_metadata_path(source, code)))


def save_kline_cache_metadata(source, code, df, *, qfq_anchor_date=None):
    normalized = normalize_kline_df(df)
    signature = _kline_cache_file_signature(source, code)
    if normalized.empty or not signature:
        return False
    anchor_date = qfq_anchor_date or _kline_anchor_date_from_frame(normalized)
    payload = {
        "version": KLINE_QFQ_CACHE_VERSION,
        "cache_version": KLINE_QFQ_CACHE_VERSION,
        "source": str(source),
        "code": str(code),
        "adjustment": "qfq",
        "qfq_anchor_date": str(anchor_date) if anchor_date else None,
        "provider_actual_end_date": str(normalized["date"].max()),
        "rows": int(len(normalized)),
        "min_date": str(normalized["date"].min()),
        "max_date": str(normalized["date"].max()),
        **signature,
    }
    _atomic_write_json(payload, kline_cache_metadata_path(source, code))
    return True


def kline_cache_metadata_matches(source, code, df, *, max_qfq_anchor_date=None):
    if df is None or df.empty:
        return False
    payload = _read_json(kline_cache_metadata_path(source, code))
    signature = _kline_cache_file_signature(source, code)
    if not signature:
        return False
    normalized = normalize_kline_df(df)
    return (
        payload.get("version") == KLINE_QFQ_CACHE_VERSION
        and payload.get("adjustment") == "qfq"
        and payload.get("source") == str(source)
        and payload.get("code") == str(code)
        and payload.get("rows") == len(normalized)
        and payload.get("min_date") == str(normalized["date"].min())
        and payload.get("max_date") == str(normalized["date"].max())
        and payload.get("size") == signature["size"]
        and payload.get("mtime_ns") == signature["mtime_ns"]
        and _kline_anchor_allowed(payload, max_qfq_anchor_date)
    )


def invalidate_kline_cache_metadata(source, code):
    try:
        os.remove(kline_cache_metadata_path(source, code))
    except FileNotFoundError:
        return False
    return True


def normalize_kline_df(df):
    if df is None or df.empty:
        return pd.DataFrame()
    df = df.copy()
    keep_cols = [
        "date",
        "code",
        "open",
        "high",
        "low",
        "close",
        "volume",
        "amount",
        "tradestatus",
    ]
    for col in keep_cols:
        if col not in df.columns:
            df[col] = np.nan
    df = df[keep_cols]
    df["date"] = pd.to_datetime(df["date"], errors="coerce").dt.strftime("%Y-%m-%d")
    for col in ["open", "high", "low", "close", "volume", "amount"]:
        df[col] = pd.to_numeric(df[col], errors="coerce")
    df = df.dropna(subset=["date", "high", "low", "close"]).sort_values("date")
    return df.drop_duplicates(subset=["date"], keep="last").reset_index(drop=True)


def combine_kline_frames(frames):
    """Combine same-source K-lines while keeping non-null enrichment fields."""
    available = [frame for frame in frames if frame is not None and not frame.empty]
    if not available:
        return pd.DataFrame()
    combined = normalize_kline_df(pd.concat(available, ignore_index=True, sort=False))
    if combined.empty:
        return combined
    return (
        combined.sort_values("date")
        .groupby("date", as_index=False)
        .last()
        .reset_index(drop=True)
    )


def load_cached_kline(source, code):
    path = kline_cache_path(source, code)
    try:
        if os.path.exists(path):
            return normalize_kline_df(pd.read_csv(path, dtype={"code": str}))
    except Exception as exc:
        print(f"{code} K线缓存读取失败: {exc}")
    return pd.DataFrame()


def load_persisted_kline(
    repository,
    source,
    code,
    start_date,
    end_date,
    *,
    max_qfq_anchor_date=None,
):
    if repository is None:
        return pd.DataFrame()
    try:
        return normalize_kline_df(
            repository.load_stock_kline(
                source,
                code,
                start_date=start_date,
                end_date=end_date,
                max_qfq_anchor_date=max_qfq_anchor_date,
            )
        )
    except Exception as exc:
        print(f"{code} DuckDB K线读取失败: {exc}")
        return pd.DataFrame()


def save_cached_kline(source, code, df):
    if df is None or df.empty:
        return 0
    path = kline_cache_path(source, code)
    tmp_path = f"{path}.{os.getpid()}.tmp"
    try:
        os.makedirs(os.path.dirname(path), exist_ok=True)
        normalized = normalize_kline_df(df)
        if normalized.empty:
            raise ValueError("K-line cache contains no valid OHLC rows")
        normalized.to_csv(tmp_path, index=False, encoding="utf-8-sig")
        with open(tmp_path, "r+b") as handle:
            os.fsync(handle.fileno())
        os.replace(tmp_path, path)
        return len(normalized)
    except Exception as exc:
        try:
            if os.path.exists(tmp_path):
                os.remove(tmp_path)
        except OSError:
            pass
        raise RuntimeError(f"{code} CSV K-line cache write failed: {exc}") from exc


def save_persisted_kline(
    repository,
    source,
    code,
    df,
    *,
    replace_range=None,
    qfq_anchor_date=None,
):
    if repository is None or df is None or df.empty:
        return 0
    try:
        normalized = normalize_kline_df(df)
        if normalized.empty:
            raise ValueError("K-line persistence contains no valid OHLC rows")
        anchor_date = qfq_anchor_date or _kline_anchor_date_from_frame(normalized)
        if replace_range is not None:
            return repository.replace_stock_kline_range(
                source,
                code,
                normalized,
                start_date=replace_range[0],
                end_date=replace_range[1],
                adjustment="qfq",
                qfq_anchor_date=anchor_date,
                cache_version=KLINE_QFQ_CACHE_VERSION,
            )
        return repository.upsert_stock_kline(
            source,
            code,
            normalized,
            adjustment="qfq",
            qfq_anchor_date=anchor_date,
            cache_version=KLINE_QFQ_CACHE_VERSION,
        )
    except Exception as exc:
        raise RuntimeError(f"{code} DuckDB K-line write failed: {exc}") from exc


def load_kline_no_trade_marker(source, code):
    path = kline_no_trade_marker_path(source, code)
    try:
        with open(path, "r", encoding="utf-8") as handle:
            payload = json.load(handle)
    except (FileNotFoundError, OSError, ValueError, TypeError):
        return {}
    if (
        str(payload.get("source", "")) != str(source)
        or str(payload.get("code", "")) != str(code)
    ):
        return {}
    observation_date = pd.to_datetime(
        payload.get("observation_date"),
        errors="coerce",
    )
    if pd.isna(observation_date):
        return {}
    return {
        "source": str(source),
        "code": str(code),
        "observation_date": observation_date.strftime("%Y-%m-%d"),
    }


def save_kline_no_trade_marker(source, code, observation_date):
    path = kline_no_trade_marker_path(source, code)
    tmp_path = f"{path}.{os.getpid()}.tmp"
    payload = {
        "version": 1,
        "source": str(source),
        "code": str(code),
        "observation_date": pd.Timestamp(observation_date).strftime("%Y-%m-%d"),
    }
    try:
        os.makedirs(os.path.dirname(path), exist_ok=True)
        with open(tmp_path, "w", encoding="utf-8") as handle:
            json.dump(payload, handle, ensure_ascii=True, indent=2, sort_keys=True)
            handle.write("\n")
        os.replace(tmp_path, path)
    except Exception as exc:
        try:
            if os.path.exists(tmp_path):
                os.remove(tmp_path)
        except OSError:
            pass
        raise RuntimeError(f"{code} no-trade marker write failed: {exc}") from exc


def _qfq_ohlc_changed(cached, fresh):
    if cached is None or cached.empty or fresh is None or fresh.empty:
        return False
    old = normalize_kline_df(cached).set_index("date")
    new = normalize_kline_df(fresh).set_index("date")
    overlap = old.index.intersection(new.index)
    if overlap.empty:
        return False
    for column in ("open", "high", "low", "close"):
        if not np.allclose(
            old.loc[overlap, column].to_numpy(dtype=float),
            new.loc[overlap, column].to_numpy(dtype=float),
            rtol=1e-6,
            atol=1e-8,
            equal_nan=True,
        ):
            return True
    return False


def _has_kline_overlap(cached, fresh):
    if cached is None or cached.empty or fresh is None or fresh.empty:
        return False
    return bool(
        set(normalize_kline_df(cached)["date"])
        & set(normalize_kline_df(fresh)["date"])
    )


def _missing_expected_trade_dates(
    cached,
    fresh,
    logical_start,
    logical_end,
    expected_trade_dates,
):
    if expected_trade_dates is None:
        return []
    expected = set()
    for item in expected_trade_dates:
        parsed = pd.to_datetime(item, errors="coerce")
        if not pd.isna(parsed):
            expected.add(parsed.strftime("%Y-%m-%d"))
    expected = {
        item for item in expected if logical_start <= item <= logical_end
    }
    available = set()
    for frame in (cached, fresh):
        if frame is not None and not frame.empty:
            available.update(normalize_kline_df(frame)["date"].tolist())
    return sorted(expected - available)


def _missing_qfq_refresh_dates(cached, incremental, refreshed, start_date, end_date):
    required = set()
    for frame in (cached, incremental):
        if frame is not None and not frame.empty:
            in_window = filter_kline_range(frame, start_date, end_date)
            required.update(in_window["date"].tolist())
    refreshed_dates = set(normalize_kline_df(refreshed)["date"].tolist())
    return sorted(required - refreshed_dates)


def _fetch_kline_range(source, code, start_date, end_date, retries, retry_delay):
    loaders = {
        "tushare": load_kline_tushare,
        "akshare": load_kline_akshare,
        "baostock": load_kline_baostock,
    }
    try:
        loader = loaders[source]
    except KeyError as exc:
        raise ValueError(f"unsupported K-line source: {source}") from exc
    return loader(
        code,
        start_date,
        end_date,
        retries=retries,
        retry_delay=retry_delay,
    )


def filter_kline_range(df, start_date, end_date):
    if df is None or df.empty:
        return pd.DataFrame()
    df = normalize_kline_df(df)
    mask = (df["date"] >= start_date) & (df["date"] <= end_date)
    return df[mask].copy().reset_index(drop=True)


def load_kline_with_cache(
    source,
    code,
    start_date,
    end_date,
    retries=4,
    retry_delay=1.5,
    repository=None,
    request_sleep=0.0,
    expected_trade_dates=None,
    observation_trade_status="unknown",
    minimum_history_rows=0,
    cache_event=None,
):
    required_history_rows = max(0, int(minimum_history_rows))
    raw_file_cached = load_cached_kline(source, code)
    file_cached = raw_file_cached
    no_trade_marker = load_kline_no_trade_marker(source, code)
    file_cache_metadata_ok = not file_cached.empty and kline_cache_metadata_matches(
        source,
        code,
        file_cached,
        max_qfq_anchor_date=end_date,
    )
    if (
        not file_cache_metadata_ok
        and not file_cached.empty
        and repository is None
        and _kline_cache_metadata_missing(source, code)
        and _legacy_kline_cache_can_infer_anchor(file_cached, end_date)
    ):
        save_kline_cache_metadata(source, code, file_cached)
        file_cache_metadata_ok = kline_cache_metadata_matches(
            source,
            code,
            file_cached,
            max_qfq_anchor_date=end_date,
        )
    if not file_cache_metadata_ok:
        file_cached = pd.DataFrame()
    if not file_cached.empty:
        file_window = filter_kline_range(file_cached, start_date, end_date)
        marker_covers_end = (
            observation_trade_status == "suspended"
            and no_trade_marker.get("observation_date") == str(end_date)
        )
        endpoints_covered = (
            file_cached["date"].min() <= start_date
            and (
                file_cached["date"].max() >= end_date
                or marker_covers_end
            )
        )
        if endpoints_covered and len(file_window) >= required_history_rows:
            if cache_event is not None:
                cache_event["complete_file_cache"] = True
            return file_window

    persisted = load_persisted_kline(
        repository,
        source,
        code,
        start_date,
        end_date,
        max_qfq_anchor_date=end_date,
    )
    if (
        repository is not None
        and file_cached.empty
        and not raw_file_cached.empty
        and _kline_cache_metadata_missing(source, code)
        and _legacy_kline_cache_can_infer_anchor(raw_file_cached, end_date)
    ):
        save_kline_cache_metadata(source, code, raw_file_cached)
        if kline_cache_metadata_matches(
            source,
            code,
            raw_file_cached,
            max_qfq_anchor_date=end_date,
        ):
            file_cached = raw_file_cached
    if (
        repository is not None
        and not file_cached.empty
        and not persisted.empty
        and _qfq_ohlc_changed(file_cached, persisted)
    ):
        file_cached = combine_kline_frames([file_cached, persisted])
        save_cached_kline(source, code, file_cached)
    if repository is not None and not file_cached.empty:
        csv_in_range = filter_kline_range(file_cached, start_date, end_date)
        csv_dates = set(csv_in_range["date"]) if not csv_in_range.empty else set()
        database_dates = set(persisted["date"]) if not persisted.empty else set()
        missing_in_database = csv_dates - database_dates
        if missing_in_database:
            invalidate_kline_cache_metadata(source, code)
        save_persisted_kline(
            repository,
            source,
            code,
            csv_in_range[csv_in_range["date"].isin(missing_in_database)],
            qfq_anchor_date=end_date,
        )
    cached = combine_kline_frames([file_cached, persisted])
    effective_expected_trade_dates = set(expected_trade_dates or [])
    if observation_trade_status == "suspended":
        effective_expected_trade_dates.discard(str(end_date))
    logical_fetch_ranges = []
    if not cached.empty:
        cached_min = cached["date"].min()
        cached_max = cached["date"].max()
        cached_window = filter_kline_range(cached, start_date, end_date)
        endpoints_covered = cached_min <= start_date and cached_max >= end_date
        history_deep_enough = len(cached_window) >= required_history_rows
        if endpoints_covered and history_deep_enough:
            if repository is not None:
                if file_cached.empty:
                    save_cached_kline(source, code, cached)
                save_kline_cache_metadata(source, code, cached)
            return filter_kline_range(cached, start_date, end_date)
        if endpoints_covered and not history_deep_enough:
            logical_fetch_ranges.append((start_date, end_date))
        elif cached_min > start_date:
            left_end = (pd.to_datetime(cached_min) - pd.DateOffset(days=1)).strftime("%Y-%m-%d")
            if start_date <= left_end:
                logical_fetch_ranges.append((start_date, left_end))
        if not endpoints_covered and cached_max < end_date:
            right_start = (pd.to_datetime(cached_max) + pd.DateOffset(days=1)).strftime("%Y-%m-%d")
            if right_start <= end_date:
                marker_matches = (
                    observation_trade_status == "suspended"
                    and no_trade_marker.get("observation_date") == str(end_date)
                )
                if not marker_matches:
                    logical_fetch_ranges.append((right_start, end_date))
    else:
        logical_fetch_ranges.append((start_date, end_date))

    for logical_start, logical_end in logical_fetch_ranges:
        request_start = logical_start
        request_end = logical_end
        if not cached.empty:
            cached_min = cached["date"].min()
            cached_max = cached["date"].max()
            if logical_start > cached_max:
                request_start = cached_max
            elif logical_end < cached_min:
                request_end = cached_min
        if request_sleep > 0:
            time.sleep(request_sleep)
        fresh = normalize_kline_df(
            _fetch_kline_range(
                source,
                code,
                request_start,
                request_end,
                retries,
                retry_delay,
            )
        )
        overlap_requested = (
            request_start < logical_start or request_end > logical_end
        )
        needs_full_refresh = _qfq_ohlc_changed(cached, fresh) or (
            overlap_requested
            and not fresh.empty
            and not _has_kline_overlap(cached, fresh)
        )
        write_no_trade_marker = False
        if needs_full_refresh:
            refresh_start = min(
                value
                for value in [
                    start_date,
                    cached["date"].min() if not cached.empty else start_date,
                    fresh["date"].min() if not fresh.empty else start_date,
                ]
                if value
            )
            refresh_end = max(
                value
                for value in [
                    end_date,
                    cached["date"].max() if not cached.empty else end_date,
                    fresh["date"].max() if not fresh.empty else end_date,
                ]
                if value
            )
            if request_sleep > 0:
                time.sleep(request_sleep)
            refreshed = normalize_kline_df(
                _fetch_kline_range(
                    source,
                    code,
                    refresh_start,
                    refresh_end,
                    retries,
                    retry_delay,
                )
            )
            if refreshed.empty:
                raise RuntimeError(
                    f"{code} QFQ adjustment changed but full-window refresh was empty"
                )
            missing_refresh_dates = _missing_qfq_refresh_dates(
                cached,
                fresh,
                refreshed,
                refresh_start,
                refresh_end,
            )
            missing_expected_dates = _missing_expected_trade_dates(
                cached,
                refreshed,
                logical_start,
                logical_end,
                effective_expected_trade_dates,
            )
            if missing_refresh_dates or missing_expected_dates:
                missing = sorted(
                    set(missing_refresh_dates) | set(missing_expected_dates)
                )
                raise RuntimeError(
                    f"{code} incomplete QFQ full-window refresh; "
                    f"missing dates: {', '.join(missing[:10])}"
                )
            if len(filter_kline_range(refreshed, start_date, end_date)) < required_history_rows:
                raise RuntimeError(
                    f"{code} incomplete QFQ full-window refresh; "
                    f"history rows below {required_history_rows}"
                )
            response_dates = set(refreshed["date"].tolist())
            if (
                observation_trade_status == "suspended"
                and logical_start <= end_date <= logical_end
                and end_date not in response_dates
            ):
                if not _has_kline_overlap(cached, refreshed):
                    raise RuntimeError(
                        f"{code} suspended observation response did not include "
                        "a cached overlap date"
                    )
                write_no_trade_marker = True
            outside_window = cached[
                (cached["date"] < refresh_start) | (cached["date"] > refresh_end)
            ]
            cached = combine_kline_frames([outside_window, refreshed])
            invalidate_kline_cache_metadata(source, code)
            save_persisted_kline(
                repository,
                source,
                code,
                refreshed,
                replace_range=(refresh_start, refresh_end),
                qfq_anchor_date=refresh_end,
            )
            save_cached_kline(source, code, cached)
            if repository is not None:
                save_kline_cache_metadata(source, code, cached)
        else:
            missing_expected_dates = _missing_expected_trade_dates(
                cached,
                fresh,
                logical_start,
                logical_end,
                effective_expected_trade_dates,
            )
            if missing_expected_dates:
                raise RuntimeError(
                    f"{code} incomplete K-line response; missing expected trade "
                    f"dates: {', '.join(missing_expected_dates[:10])}"
                )
            response_dates = set(fresh["date"].tolist()) if not fresh.empty else set()
            if (
                observation_trade_status == "suspended"
                and logical_start <= end_date <= logical_end
                and end_date not in response_dates
            ):
                if not _has_kline_overlap(cached, fresh):
                    raise RuntimeError(
                        f"{code} suspended observation response did not include "
                        "a cached overlap date"
                    )
                write_no_trade_marker = True
        if not needs_full_refresh and not fresh.empty:
            candidate = combine_kline_frames([cached, fresh])
            if len(filter_kline_range(candidate, start_date, end_date)) < required_history_rows:
                raise RuntimeError(
                    f"{code} incomplete K-line response; "
                    f"history rows below {required_history_rows}"
                )
            cached = candidate
            save_cached_kline(source, code, cached)
            save_persisted_kline(
                repository,
                source,
                code,
                fresh,
                qfq_anchor_date=_kline_anchor_date_from_frame(fresh),
            )
            if repository is not None:
                save_kline_cache_metadata(source, code, cached)
        if write_no_trade_marker:
            save_kline_no_trade_marker(source, code, end_date)
    result = filter_kline_range(cached, start_date, end_date)
    if len(result) < required_history_rows:
        raise RuntimeError(
            f"{code} incomplete K-line response; "
            f"history rows below {required_history_rows}"
        )
    if repository is not None and not cached.empty:
        if file_cached.empty:
            save_cached_kline(source, code, cached)
        save_kline_cache_metadata(source, code, cached)
    return result


def load_kline_baostock(code, start_date, end_date, retries=4, retry_delay=1.5):
    fields = "date,code,open,high,low,close,volume,tradestatus"
    def query():
        return bs.ensure_success(
            bs.query_history_k_data_plus(
                code,
                fields,
                start_date=start_date,
                end_date=end_date,
                frequency="d",
                adjustflag=ADJUST_FLAG_QFQ,
            ),
            f"{code} BaoStock K-line",
        )

    rs = call_with_backoff(
        query,
        f"{code} baostock K线",
        retries=retries,
        retry_delay=retry_delay,
        on_retry=lambda _exc, _attempt: bs.reconnect(),
    )
    df = rs.get_data()
    if df.empty:
        return pd.DataFrame()
    df.columns = rs.fields
    if "tradestatus" in df.columns:
        df = df[df["tradestatus"] == "1"]
    return df


def _to_tushare_code(code):
    pure = pure_code(code)
    if "." in str(code):
        market = str(code).split(".", 1)[0].upper()
    else:
        market = "SH" if pure.startswith(("6", "9")) else "SZ"
    return f"{pure}.{market}"


def load_kline_tushare(code, start_date, end_date, retries=4, retry_delay=1.5):
    """Load and locally calculate end-date anchored Tushare QFQ prices."""
    ts_code = _to_tushare_code(code)
    params = {
        "ts_code": ts_code,
        "start_date": str(start_date).replace("-", ""),
        "end_date": str(end_date).replace("-", ""),
        "retries": retries,
        "retry_delay": retry_delay,
    }
    daily = ts_api.query(
        "daily",
        fields="ts_code,trade_date,open,high,low,close,vol,amount",
        **params,
    )
    factors = ts_api.query(
        "adj_factor",
        fields="ts_code,trade_date,adj_factor",
        **params,
    )
    if daily.empty or factors.empty:
        return pd.DataFrame()
    merged = daily.merge(factors, on=["ts_code", "trade_date"], how="inner")
    if merged.empty:
        raise RuntimeError(f"{ts_code} Tushare 日线与复权因子没有重叠日期")
    if "amount" not in merged:
        merged["amount"] = pd.NA
    for column in ["open", "high", "low", "close", "vol", "amount", "adj_factor"]:
        merged[column] = pd.to_numeric(merged[column], errors="coerce")
    merged = merged.dropna(subset=["trade_date", "open", "high", "low", "close", "adj_factor"])
    if merged.empty:
        return pd.DataFrame()
    anchor = merged.sort_values("trade_date")["adj_factor"].iloc[-1]
    if not anchor:
        raise RuntimeError(f"{ts_code} Tushare 复权因子无效")
    ratio = merged["adj_factor"] / float(anchor)
    for column in ["open", "high", "low", "close"]:
        merged[column] = merged[column] * ratio
    return pd.DataFrame({
        "date": pd.to_datetime(merged["trade_date"], errors="coerce").dt.strftime("%Y-%m-%d"),
        "code": code,
        "open": merged["open"],
        "high": merged["high"],
        "low": merged["low"],
        "close": merged["close"],
        "volume": merged["vol"],
        "amount": merged["amount"] * 1000.0,
    }).sort_values("date").reset_index(drop=True)


def resolve_price_source(requested, start_date, end_date, retries, retry_delay):
    if requested != "tushare":
        return requested
    if not get_tushare_token():
        raise RuntimeError("--price-source tushare 需要配置 Tushare token")
    print("前复权 K 线来源: Tushare（显式小批量模式）")
    return "tushare"


def _normalize_akshare_kline_columns(df):
    """Map AkShare K-line columns while preserving provider OHLC values."""
    if df is None or df.empty:
        return pd.DataFrame()
    aliases = {
        "\u65e5\u671f": "date",
        "\u80a1\u7968\u4ee3\u7801": "code",
        "\u5f00\u76d8": "open",
        "\u6536\u76d8": "close",
        "\u6700\u9ad8": "high",
        "\u6700\u4f4e": "low",
        "\u6210\u4ea4\u91cf": "volume",
        "\u6210\u4ea4\u989d": "amount",
        "\u6362\u624b\u7387": "turnover",
        "date": "date",
        "code": "code",
        "open": "open",
        "close": "close",
        "high": "high",
        "low": "low",
        "volume": "volume",
        "amount": "amount",
        "turnover": "turnover",
    }
    return df.rename(
        columns={
            column: aliases[str(column).strip()]
            for column in df.columns
            if str(column).strip() in aliases
        }
    ).copy()


def load_kline_akshare(code, start_date, end_date, retries=4, retry_delay=1.5):
    pure = pure_code(code)
    if "." in str(code):
        market = str(code).split(".", 1)[0].lower()
    else:
        market = "sh" if pure.startswith(("6", "9")) else "sz"
    daily_symbol = f"{market}{pure}"

    if pure.startswith("689"):
        df = call_with_backoff(
            lambda: ak.stock_zh_a_hist_tx(
                symbol=daily_symbol,
                start_date=str(start_date).replace("-", ""),
                end_date=str(end_date).replace("-", ""),
                adjust="qfq",
                timeout=15,
            ),
            f"{code} akshare腾讯CDR前复权K线",
            retries=retries,
            retry_delay=retry_delay,
        )
        if df is None or df.empty:
            return pd.DataFrame()
        df = df.copy()
        df["code"] = code
        if "volume" not in df.columns and "amount" in df.columns:
            df = df.rename(columns={"amount": "volume"})
        return df

    def fetch_daily():
        result = ak.stock_zh_a_hist(
            symbol=pure,
            period="daily",
            start_date=str(start_date).replace("-", ""),
            end_date=str(end_date).replace("-", ""),
            adjust="qfq",
        )
        if result is None or result.empty:
            raise RuntimeError("AkShare Sina K-line returned an empty result")
        return result

    try:
        df = call_with_backoff(
            fetch_daily,
            f"{code} akshare东方财富K线",
            retries=retries,
            retry_delay=retry_delay,
        )
    except Exception as exc:
        print(f"{code} akshare东方财富K线失败，回退新浪: {exc}")
        df = call_with_backoff(
            lambda: ak.stock_zh_a_daily(
                symbol=daily_symbol,
                start_date=str(start_date).replace("-", ""),
                end_date=str(end_date).replace("-", ""),
                adjust="qfq",
            ),
            f"{code} akshare新浪K线",
            retries=retries,
            retry_delay=retry_delay,
        )
    if df is None or df.empty:
        return pd.DataFrame()
    if "date" in df.columns:
        df = df.copy()
        df["code"] = code
        return df.rename(columns={
            "date": "date",
            "open": "open",
            "high": "high",
            "low": "low",
            "close": "close",
            "volume": "volume",
            "amount": "amount",
        })
    return df.rename(columns={
        "日期": "date",
        "股票代码": "code",
        "开盘": "open",
        "最高": "high",
        "最低": "low",
        "收盘": "close",
        "成交量": "volume",
    })


def update_fetch_progress(counts, result):
    """Update human-readable stock-fetch outcome counters from one task result."""
    counts["processed"] += 1
    if not result:
        counts["skipped"] += 1
        return
    status = next(
        (item for item in result if item.get("signal_type") == "STATUS"),
        {},
    )
    if status.get("observation_status") == "data_unavailable":
        counts["failed"] += 1
    else:
        counts["succeeded"] += 1
    if status.get("data_origin") == "complete_file_cache":
        counts["cache_hits"] += 1
    counts["signals"] += sum(
        item.get("signal_type") in {"BASE", "XG"} for item in result
    )


def format_fetch_progress(counts, total):
    remaining = max(0, int(total) - int(counts["processed"]))
    return (
        f"行情进度 {counts['processed']}/{total} | 成功 {counts['succeeded']} | "
        f"完整缓存 {counts['cache_hits']} | 跳过 {counts['skipped']} | "
        f"失败 {counts['failed']} | 剩余 {remaining} | "
        f"BASE/XG记录 {counts['signals']}"
    )


def fetch_one_stock(task):
    if len(task) == 16:
        task = (*task, False, "unknown")
    elif len(task) == 17:
        task = (*task, "unknown")
    elif len(task) != 18:
        raise ValueError(f"fetch task must contain 16, 17, or 18 values, got {len(task)}")
    (
        code,
        name,
        start_date,
        end_date,
        date_set,
        mktcap_yi,
        ipo_date,
        min_mktcap,
        min_list_days,
        sleep,
        price_source,
        retries,
        retry_delay,
        debug_filters,
        require_end_trade,
        missing_mktcap_policy,
        persist_kline,
        observation_trade_status,
    ) = task
    if ipo_date is None:
        return []
    ipo_ts = pd.to_datetime(ipo_date, errors="coerce")
    if pd.isna(ipo_ts):
        return []
    end_ts = pd.to_datetime(end_date, errors="coerce")
    if pd.isna(end_ts):
        raise ValueError(f"invalid K-line end date: {end_date}")
    current_list_days = (end_ts - ipo_ts).days
    if current_list_days <= min_list_days:
        return []
    requested_start = pd.to_datetime(start_date, errors="coerce")
    if pd.isna(requested_start):
        raise ValueError(f"invalid K-line start date: {start_date}")
    effective_start_date = max(requested_start, ipo_ts).strftime("%Y-%m-%d")
    eligible_trade_dates = [
        value
        for value in date_set
        if pd.to_datetime(value, errors="coerce") >= ipo_ts
    ]
    minimum_history_rows = min(
        30,
        max(1, int(len(eligible_trade_dates) * 0.8)),
    )
    mktcap_missing_ok = (
        min_mktcap is not None
        and mktcap_yi is None
        and missing_mktcap_policy == "pass"
    )
    fetch_error = ""
    kline_repository = KlineRepository(Database()) if persist_kline else None
    cache_event = {}
    try:
        df = load_kline_with_cache(
            price_source,
            code,
            effective_start_date,
            end_date,
            retries=retries,
            retry_delay=retry_delay,
            repository=kline_repository,
            request_sleep=sleep,
            expected_trade_dates={end_date},
            observation_trade_status=observation_trade_status,
            minimum_history_rows=minimum_history_rows,
            cache_event=cache_event,
        )
    except Exception as exc:
        fetch_error = str(exc)
        df = load_cached_kline(price_source, code)
    data_origin = (
        "complete_file_cache"
        if cache_event.get("complete_file_cache")
        else "database_or_network"
    )
    if df.empty:
        return [{
            "signal_type": "STATUS",
            "code": code,
            "name": name,
            "latest_data_date": "",
            "observation_status": "data_unavailable",
            "error": fetch_error or "empty kline",
            "data_origin": data_origin,
        }]
    for col in ["open", "high", "low", "close", "volume"]:
        df[col] = pd.to_numeric(df[col], errors="coerce")
    df["date"] = pd.to_datetime(df["date"], errors="coerce").dt.strftime("%Y-%m-%d")
    df = df.dropna(subset=["date", "high", "low", "close"]).sort_values("date").reset_index(drop=True)
    latest_data_date = df["date"].max() if not df.empty else ""
    observation_status = classify_observation_status(
        latest_data_date,
        end_date,
        fetch_error=fetch_error,
    )
    if len(df) < 30:
        return [{
            "signal_type": "STATUS",
            "code": code,
            "name": name,
            "latest_data_date": latest_data_date,
            "observation_status": "data_unavailable",
            "error": fetch_error or "insufficient kline history",
            "data_origin": data_origin,
        }]

    k = calc_kdj_k(df)
    wr10 = calc_wr(df, 10)
    wr20 = calc_wr(df, 20)
    rsi9 = calc_rsi(df["close"], 9)
    current_mktcap_ok = bool(
        min_mktcap is not None
        and (
            mktcap_missing_ok
            or (
                mktcap_yi is not None
                and not pd.isna(mktcap_yi)
                and float(mktcap_yi) > float(min_mktcap)
            )
        )
    )
    list_days = (
        pd.to_datetime(df["date"], errors="coerce") - ipo_ts
    ).dt.days
    list_days_ok = list_days > min_list_days

    kd80 = k > 80
    wr3 = (wr10 < 20) & (wr20 < 20)
    rsi70 = rsi9 > 70
    base = kd80 & wr3 & rsi70 & list_days_ok
    xg = base.rolling(5, min_periods=5).sum() == 5
    market_cap_base = base & current_mktcap_ok
    market_cap_xg = market_cap_base.rolling(5, min_periods=5).sum() == 5

    hits = []
    debug_rows = []
    for idx, row in df.iterrows():
        row = df.loc[idx]
        if row["date"] not in date_set:
            continue
        row_date = pd.to_datetime(row["date"], errors="coerce")
        row_list_days = (
            (row_date - ipo_ts).days if pd.notna(row_date) else current_list_days
        )
        if min_mktcap is None or mktcap_yi is None or pd.isna(mktcap_yi):
            mktcap_at_date = np.nan
        else:
            mktcap_at_date = float(mktcap_yi)
        if debug_filters:
            debug_rows.append({
                "date": row["date"],
                "code": code,
                "name": name,
                "signal_type": "DEBUG",
                "kd80": bool(kd80.loc[idx]),
                "wr3": bool(wr3.loc[idx]),
                "rsi70": bool(rsi70.loc[idx]),
                "mktcap_ok": bool(current_mktcap_ok),
                "list_days_ok": bool(list_days_ok.loc[idx]),
                "base_ok": bool(base.loc[idx]),
                "xg_ok": bool(xg.loc[idx]),
                "market_cap_ok": bool(current_mktcap_ok),
                "market_cap_base_ok": bool(market_cap_base.loc[idx]),
                "market_cap_xg_ok": bool(market_cap_xg.loc[idx]),
            })
        if not bool(list_days_ok.loc[idx]):
            continue
        record = {
            "date": row["date"],
            "code": code,
            "name": name,
            "close": row["close"],
            "mktcap_yi": round(mktcap_at_date, 2) if pd.notna(mktcap_at_date) else np.nan,
            "list_days": int(row_list_days),
            "kdj_k": round(float(k.loc[idx]), 2),
            "wr10": round(float(wr10.loc[idx]), 2),
            "wr20": round(float(wr20.loc[idx]), 2),
            "rsi9": round(float(rsi9.loc[idx]), 2),
            "market_cap_ok": bool(current_mktcap_ok),
        }
        if bool(base.loc[idx]):
            base_record = record.copy()
            base_record["signal_type"] = "BASE"
            hits.append(base_record)
        if bool(xg.loc[idx]):
            xg_record = record.copy()
            xg_record["signal_type"] = "XG"
            hits.append(xg_record)
        if bool(market_cap_base.loc[idx]):
            market_cap_base_record = record.copy()
            market_cap_base_record["signal_type"] = "MARKET_CAP_BASE"
            hits.append(market_cap_base_record)
        if bool(market_cap_xg.loc[idx]):
            market_cap_xg_record = record.copy()
            market_cap_xg_record["signal_type"] = "MARKET_CAP_XG"
            hits.append(market_cap_xg_record)
    coverage_row = {
        "signal_type": "TRADE_COVERAGE",
        "code": code,
        "covered_dates": tuple(
            df.loc[df["date"].isin(date_set), "date"].drop_duplicates()
        ),
    }
    status_row = {
        "signal_type": "STATUS",
        "code": code,
        "name": name,
        "latest_data_date": latest_data_date,
        "observation_status": observation_status,
        "error": fetch_error,
        "data_origin": data_origin,
    }
    return hits + debug_rows + [coverage_row, status_row]


def calc_streaks(counts):
    rows = []
    up_streak = 0
    down_streak = 0
    prev = None
    for date, count in counts:
        change = 0 if prev is None else count - prev
        if prev is None or change == 0:
            up_streak = 0
            down_streak = 0
        elif change > 0:
            up_streak += 1
            down_streak = 0
        else:
            down_streak += 1
            up_streak = 0
        if up_streak >= 5:
            signal = "结构转好确认，右侧成功率提升"
        elif up_streak >= 3:
            signal = "结构初步转好"
        elif down_streak >= 5:
            signal = "结构转坏确认，右侧成功率下降"
        elif down_streak >= 3:
            signal = "结构初步转坏"
        else:
            signal = "观察"
        rows.append({
            "date": date,
            "count": count,
            "change": change,
            "up_streak": up_streak,
            "down_streak": down_streak,
            "signal": signal,
        })
        prev = count
    return pd.DataFrame(rows)


def build_formula_summary(
    hits,
    trade_dates,
    output_days=21,
    trade_coverage=None,
    current_statuses=None,
):
    """Combine compatibility daily counts with rolling breadth statistics."""
    dates = [str(value) for value in trade_dates]
    if hits is None or hits.empty or "signal_type" not in hits.columns:
        xg_hits = pd.DataFrame(columns=["date", "code"])
        base_counts = {}
        xg_counts = {}
    else:
        xg_hits = hits[hits["signal_type"] == "XG"].copy()
        base_hits = hits[hits["signal_type"] == "BASE"].copy()
        xg_hits["date"] = pd.to_datetime(
            xg_hits["date"], errors="coerce",
        ).dt.strftime("%Y-%m-%d")
        base_hits["date"] = pd.to_datetime(
            base_hits["date"], errors="coerce",
        ).dt.strftime("%Y-%m-%d")
        xg_counts = xg_hits.groupby("date").size().to_dict()
        base_counts = base_hits.groupby("date").size().to_dict()

    compatibility = calc_streaks(
        [(date, int(xg_counts.get(date, 0))) for date in dates]
    )
    compatibility["base_count"] = compatibility["date"].map(
        lambda date: int(base_counts.get(date, 0))
    )
    compatibility["date"] = pd.to_datetime(
        compatibility["date"], errors="coerce",
    ).dt.strftime("%Y-%m-%d")
    compatibility = compatibility.tail(output_days).reset_index(drop=True)
    rolling = build_window_trend(
        xg_hits,
        dates,
        window=21,
        output_days=output_days,
        trade_coverage=trade_coverage,
        current_statuses=current_statuses,
    )
    summary = compatibility.merge(rolling, on="date", how="left")
    zero_columns = [
        "window_unique_count",
        "technical_unique_count",
        "tradable_unique_count",
        "window_change",
        "window_up_streak",
        "window_down_streak",
        "trend_up_streak",
        "trend_down_streak",
    ]
    for column in zero_columns:
        summary[column] = pd.to_numeric(summary[column], errors="coerce").fillna(0).astype(int)
    summary["window_trend_slope"] = pd.to_numeric(
        summary["window_trend_slope"], errors="coerce",
    ).fillna(0.0)
    summary["trend_signal"] = summary["trend_signal"].fillna("neutral")
    return summary[
        [
            "date",
            "base_count",
            "count",
            "change",
            "up_streak",
            "down_streak",
            "signal",
            "window_unique_count",
            "technical_unique_count",
            "tradable_unique_count",
            "window_change",
            "window_up_streak",
            "window_down_streak",
            "window_trend_slope",
            "trend_up_streak",
            "trend_down_streak",
            "trend_signal",
        ]
    ]


FORMULA_EXCEL_HEADERS = {
    "date": "日期",
    "base_count": "当日BASE数量",
    "count": "当日XG数量",
    "change": "较前一日变化",
    "up_streak": "连续增加天数",
    "down_streak": "连续减少天数",
    "signal": "当日结构判断",
    "window_unique_count": "近21日正式可交易去重数",
    "technical_unique_count": "近21日技术命中去重数",
    "tradable_unique_count": "观察日可交易去重数",
    "window_change": "较前一观察节点变化",
    "window_up_streak": "窗口连续增加节点",
    "window_down_streak": "窗口连续减少节点",
    "window_trend_slope": "窗口趋势斜率",
    "trend_up_streak": "斜率连续为正节点",
    "trend_down_streak": "斜率连续为负节点",
    "trend_signal": "21日趋势判断",
    "market_cap_unique_count": "市值超100亿正式池",
    "market_cap_technical_unique_count": "市值超100亿技术池",
    "suspended_count": "观察日停牌技术命中",
    "unavailable_count": "数据不可用数量",
    "signal_type": "信号类型",
    "code": "股票代码",
    "name": "股票名称",
    "close": "命中日收盘价",
    "mktcap_yi": "总市值（亿元）",
    "list_days": "上市天数",
    "kdj_k": "KDJ-K",
    "wr10": "WR10",
    "wr20": "WR20",
    "rsi9": "RSI9",
    "latest_data_date": "最新行情日期",
    "observation_status": "观察日状态",
    "error": "异常说明",
    "data_origin": "行情来源状态",
}

FORMULA_EXCEL_VALUES = {
    "BASE": "BASE（当日四条件同时满足）",
    "XG": "XG（BASE连续5日满足）",
    "MARKET_CAP_BASE": "市值池BASE",
    "MARKET_CAP_XG": "市值池XG",
    "tradable": "可交易",
    "suspended_or_no_trade": "观察日停牌或无交易",
    "data_unavailable": "数据不可用",
    "complete_file_cache": "完整CSV缓存命中",
    "database_or_network": "DuckDB或网络补齐",
}


def _excel_display_value(column, value):
    if column in {"signal_type", "observation_status", "data_origin"}:
        return FORMULA_EXCEL_VALUES.get(value, value)
    return value


def _append_excel_table(sheet, frame, columns):
    sheet.append([FORMULA_EXCEL_HEADERS.get(column, column) for column in columns])
    if frame is None or frame.empty:
        return
    for row in frame[columns].to_dict("records"):
        sheet.append(
            [_excel_display_value(column, row.get(column)) for column in columns]
        )


def save_workbook(
    summary,
    hits,
    sample=False,
    unique_hits=None,
    technical_unique_hits=None,
    market_cap_unique_hits=None,
    suspended_technical_hits=None,
    statuses=None,
):
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    suffix = "_sample" if sample else ""
    path = os.path.join(OUTPUT_DIR, f"formula33_stats_{stamp}{suffix}.xlsx")
    csv_path = os.path.join(OUTPUT_DIR, f"formula33_stats_{stamp}{suffix}.csv")
    summary.to_csv(csv_path, index=False, encoding="utf-8-sig")

    wb = Workbook()
    guide = wb.active
    guide.title = "先看这里"
    guide_rows = [
        ["阅读顺序", "先看“近21日正式名单”，再看“33公式日统计”和“观察日状态”；诊断表不等于正式候选。"],
        ["正式结果是什么", "近21个交易日内至少一次满足XG，且观察日仍可交易的股票去重名单。"],
        ["BASE", "某一交易日同时满足 KDJ-K>80、WR10/WR20<20、RSI9>70、上市超过300天。"],
        ["XG", "BASE连续5个交易日成立；这是正式技术信号。"],
        ["近21日技术命中", "只看技术条件，不检查观察日是否停牌；用于解释差异，不直接作为正式名单。"],
        ["市值超100亿池", "从技术命中中单独列出的总市值参考池，不改变原始XG定义。"],
        ["观察日状态", "可交易=行情覆盖观察日；停牌或无交易=仅进诊断；数据不可用=接口或历史不足，需要处理。"],
        ["指标方向", "KDJ-K、RSI9越高代表短期偏强；WR10/WR20越低代表收盘越接近近期高位。"],
        ["机器数据", "同名CSV保留英文稳定字段，供程序读取；Excel使用中文标题供人工查看。"],
    ]
    for row in guide_rows:
        guide.append(row)

    ws = wb.create_sheet("33公式日统计")
    headers = ["date", "base_count", "count", "change", "up_streak", "down_streak", "signal"]
    headers += [
        column
        for column in [
            "window_unique_count",
            "window_change",
            "window_up_streak",
            "window_down_streak",
            "window_trend_slope",
            "trend_up_streak",
            "trend_down_streak",
            "trend_signal",
            "technical_unique_count",
            "tradable_unique_count",
            "market_cap_unique_count",
            "market_cap_technical_unique_count",
            "suspended_count",
            "unavailable_count",
        ]
        if column in summary.columns
    ]
    _append_excel_table(ws, summary, headers)

    ws2 = wb.create_sheet("横向统计")
    ws2.append(["指标"] + summary["date"].tolist())
    ws2.append(["BASE数量"] + summary.get("base_count", pd.Series([np.nan] * len(summary))).tolist())
    ws2.append(["XG数量"] + summary["count"].tolist())
    ws2.append(["较前日变化"] + summary["change"].tolist())
    ws2.append(["连续上升"] + summary["up_streak"].tolist())
    ws2.append(["连续下降"] + summary["down_streak"].tolist())
    ws2.append(["结构信号"] + summary["signal"].tolist())
    if "window_unique_count" in summary.columns:
        ws2.append(["21日XG可交易技术去重"] + summary["window_unique_count"].tolist())
        ws2.append(["21日节点较前节点变化"] + summary["window_change"].tolist())
        ws2.append(["21日节点连续上升"] + summary["window_up_streak"].tolist())
        ws2.append(["21日节点连续下降"] + summary["window_down_streak"].tolist())
        if "technical_unique_count" in summary.columns:
            ws2.append(["21日XG技术全量"] + summary["technical_unique_count"].tolist())
        if "market_cap_unique_count" in summary.columns:
            ws2.append(["21日市值大于100亿池"] + summary["market_cap_unique_count"].tolist())
        ws2.append(["21日节点回归趋势斜率"] + summary["window_trend_slope"].tolist())
        ws2.append(["斜率连续为正"] + summary["trend_up_streak"].tolist())
        ws2.append(["斜率连续为负"] + summary["trend_down_streak"].tolist())

    ws3 = wb.create_sheet("命中股票")
    cols = ["signal_type", "date", "code", "name", "close", "mktcap_yi", "list_days", "kdj_k", "wr10", "wr20", "rsi9"]
    _append_excel_table(ws3, hits, cols)

    ws4 = wb.create_sheet("近21日正式名单")
    unique_cols = ["date", "code", "name", "close", "mktcap_yi", "list_days", "kdj_k", "wr10", "wr20", "rsi9"]
    _append_excel_table(ws4, unique_hits, unique_cols)

    ws5 = wb.create_sheet("近21日技术命中")
    _append_excel_table(ws5, technical_unique_hits, unique_cols)

    ws6 = wb.create_sheet("市值大于100亿池")
    _append_excel_table(ws6, market_cap_unique_hits, unique_cols)

    ws7 = wb.create_sheet("停牌技术命中诊断")
    _append_excel_table(ws7, suspended_technical_hits, unique_cols)

    ws8 = wb.create_sheet("观察日状态")
    status_cols = ["code", "name", "latest_data_date", "observation_status", "data_origin", "error"]
    if statuses is not None and "data_origin" not in statuses.columns:
        statuses = statuses.copy()
        statuses["data_origin"] = ""
    _append_excel_table(ws8, statuses, status_cols)

    yellow = PatternFill("solid", fgColor="FFF2CC")
    green = PatternFill("solid", fgColor="E2F0D9")
    red = PatternFill("solid", fgColor="F4CCCC")
    thin = Side(style="thin", color="666666")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for sheet in wb.worksheets:
        sheet.freeze_panes = "A2"
        for row in sheet.iter_rows():
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                if cell.row == 1:
                    cell.font = Font(bold=True)
                    cell.fill = yellow
                if isinstance(cell.value, str) and "转好" in cell.value:
                    cell.fill = green
                if isinstance(cell.value, str) and "转坏" in cell.value:
                    cell.fill = red
        for col in range(1, sheet.max_column + 1):
            sheet.column_dimensions[get_column_letter(col)].width = 18
        if sheet.title != "先看这里" and sheet.max_row >= 1:
            sheet.auto_filter.ref = sheet.dimensions
    guide.column_dimensions["A"].width = 22
    guide.column_dimensions["B"].width = 90
    wb.save(path)
    return path, csv_path


def build_sample(lookback):
    dates = pd.bdate_range(end=pd.Timestamp("2026-06-24"), periods=lookback).strftime("%Y-%m-%d").tolist()
    counts = [42, 45, 48, 52, 57, 60, 58, 55, 51, 48, 46, 49, 53, 58, 64, 70, 74, 72, 69, 66, 63][-lookback:]
    summary = calc_streaks(list(zip(dates, counts)))
    summary["base_count"] = [value + 80 for value in counts]
    summary = summary[["date", "base_count", "count", "change", "up_streak", "down_streak", "signal"]]
    hit_rows = []
    for date, count in zip(dates, counts):
        for idx in range(min(count, 12)):
            hit_rows.append({
                "signal_type": "XG",
                "date": date,
                "code": f"sz.30{idx:04d}",
                "name": f"33样本{idx + 1}",
                "close": 10 + idx,
                "mktcap_yi": 120 + idx * 5,
                "list_days": 500 + idx,
                "kdj_k": 82 + idx % 10,
                "wr10": 8 + idx % 8,
                "wr20": 9 + idx % 7,
                "rsi9": 72 + idx % 12,
            })
    return summary, pd.DataFrame(hit_rows)


def main(argv=None):
    args = parse_args(argv)
    if args.sample:
        summary, hits = build_sample(args.lookback)
        xlsx_path, csv_path = save_workbook(summary, hits, sample=True)
        print(f"Excel已保存: {xlsx_path}")
        print(f"CSV已保存: {csv_path}")
        print(summary.tail(8).to_string(index=False))
        return

    if reuse_completed_manifest_without_network(args):
        return 0

    bs_available = False
    if args.price_source == "baostock" or args.metadata_source in ("baostock", "auto"):
        lg = bs.login()
        bs_available = lg.error_code == "0"
        if not bs_available:
            print(f"Baostock不可用，改用 akshare 元数据: {lg.error_msg}")
            if args.price_source == "baostock":
                raise SystemExit("Baostock不可用时不能使用 --price-source baostock，请改用 --price-source akshare")
    try:
        calculation_days = args.lookback * 3 - 2
        effective_end_date = resolve_auto_end_date(
            args.end_date, args.data_ready_time
        )
        calendar_extra_days = args.history_days + 90
        if args.start_date:
            requested_start = pd.to_datetime(args.start_date, errors="coerce")
            requested_end = pd.to_datetime(effective_end_date, errors="coerce")
            if pd.notna(requested_start) and pd.notna(requested_end):
                calendar_extra_days = max(
                    calendar_extra_days,
                    int((requested_end.normalize() - requested_start.normalize()).days) + 30,
                )
        raw_trade_dates = (
            get_trade_dates(calculation_days + 5, calendar_extra_days)
            if bs_available and args.metadata_source in ("baostock", "auto")
            else []
        )
        if len(raw_trade_dates) < calculation_days:
            raw_trade_dates = get_trade_dates_akshare(
                calculation_days + 5,
                calendar_extra_days,
                required_through=effective_end_date,
            )
        if not raw_trade_dates or (
            not args.start_date and len(raw_trade_dates) < calculation_days
        ):
            raise SystemExit("交易日不足，无法统计")
        if not args.end_date:
            print(f"未指定 --end-date，按数据可用时间 {args.data_ready_time} 使用截止日: {effective_end_date}")
        trade_dates = select_trade_dates(
            raw_trade_dates,
            args.start_date,
            effective_end_date,
            calculation_days,
        )
        if not trade_dates or (
            not args.start_date and len(trade_dates) < calculation_days
        ):
            raise SystemExit(
                f"交易日不足，无法计算最近 {args.lookback} 个交易日的滚动趋势"
            )
        output_dates = trade_dates[-args.lookback:]
        latest_date = trade_dates[-1]
        completion_manifest = CompletionManifest(FORMULA33_MANIFEST_FILE)
        completion_arguments = build_completion_arguments(args, latest_date)
        cached_universe = load_cached_universe()
        if not cached_universe.empty:
            cached_universe = select_universe_for_run(cached_universe, args)
            if completion_manifest.matches(
                observation_date=latest_date,
                arguments=completion_arguments,
                universe_codes=cached_universe["code"].astype(str).tolist(),
                code_version=FORMULA33_CODE_VERSION,
            ):
                print(
                    f"Formula33 resume: completed manifest hit date={latest_date}; "
                    "network_fetch=0"
                )
                return 0
        if bs_available and args.metadata_source == "baostock":
            universe_date, universe = get_universe_with_fallback(trade_dates)
        else:
            universe_date, universe = latest_date, load_universe_snapshot(latest_date)
        if universe.empty:
            raise SystemExit("无法获取沪深A股股票池")
        if universe_date != latest_date:
            print(f"股票池使用 {universe_date}，统计交易日仍使用 {latest_date} 之前最近 {len(trade_dates)} 个交易日")
            trade_dates = select_trade_dates(
                raw_trade_dates,
                args.start_date,
                universe_date,
                calculation_days,
            )
            if not trade_dates or (
                not args.start_date and len(trade_dates) < calculation_days
            ):
                raise SystemExit(
                    f"股票池最新日期为 {universe_date}，但滚动趋势交易日不足"
                )
            output_dates = trade_dates[-args.lookback:]
            latest_date = trade_dates[-1]
            print(f"已按股票池最新日期裁剪统计窗口，最新统计日: {latest_date}")
        print(f"本次统计区间: {trade_dates[0]} ~ {trade_dates[-1]}，共 {len(trade_dates)} 个交易日")
        start_date = (pd.to_datetime(trade_dates[0]) - pd.DateOffset(days=args.history_days)).strftime("%Y-%m-%d")
        universe = select_universe_for_run(universe, args, announce=True)
        completion_arguments = build_completion_arguments(args, latest_date)
        universe_codes = universe["code"].astype(str).tolist()
        if completion_manifest.matches(
            observation_date=latest_date,
            arguments=completion_arguments,
            universe_codes=universe_codes,
            code_version=FORMULA33_CODE_VERSION,
        ):
            print(
                f"Formula33 resume: completed manifest hit date={latest_date}; "
                "network_fetch=0"
            )
            return 0
        price_source = resolve_price_source(
            args.price_source,
            start_date,
            latest_date,
            args.retries,
            args.retry_delay,
        )
        Database().initialize()
        print("DuckDB K线持久化已启用：每只股票查到后立即写入 raw.stock_kline_daily，并保留CSV缓存。")
        basic = load_stock_basic() if bs_available and args.metadata_source == "baostock" else pd.DataFrame()
        if basic.empty:
            basic = load_stock_basic_snapshot(latest_date, universe)
        list_date_map = {}
        if not basic.empty and "code" in basic.columns and "ipoDate" in basic.columns:
            list_date_map = dict(zip(basic["code"], basic["ipoDate"]))
        list_date_coverage = require_lookup_coverage(
            "上市日期",
            universe,
            list_date_map,
            minimum=MIN_LIST_DATE_COVERAGE,
        )
        latest_calendar_date = latest_trade_date_from_calendar_snapshot()
        spot_snapshot = load_observation_spot_snapshot(
            latest_date,
            allow_network=bool(latest_calendar_date)
            and latest_date == latest_calendar_date,
        )
        observation_trade_status, observation_status_coverage = (
            build_observation_trade_status(
                universe,
                spot_snapshot,
                minimum=MIN_OBSERVATION_STATUS_COVERAGE,
            )
            if not spot_snapshot.empty
            else ({code: "unknown" for code in universe_codes}, 0.0)
        )
        suspended_codes = {
            code
            for code, status in observation_trade_status.items()
            if status == "suspended"
        }
        print(
            "观察日交易状态: "
            f"coverage={observation_status_coverage:.1%} "
            f"confirmed_suspended={len(suspended_codes)}"
        )
        try:
            cap_map, cap_source = load_market_cap_snapshot(
                args.market_cap_source,
                latest_date,
                universe,
                capital_workers=args.capital_workers,
                capital_sleep=args.capital_sleep,
                retries=args.retries,
                retry_delay=args.retry_delay,
                spot_snapshot=spot_snapshot,
            )
        except Exception as exc:
            raise SystemExit(f"{exc}\n该公式需要 FINANCE(40)>100亿；可配置 Tushare token，或网络可用时使用 akshare。")
        min_mktcap = None if cap_source == "none" else args.min_mktcap
        market_cap_coverage = (
            1.0
            if cap_source == "none"
            else require_lookup_coverage(
                "总市值",
                universe,
                cap_map,
                minimum=MIN_MARKET_CAP_COVERAGE,
            )
        )

        date_set = set(trade_dates)
        tasks = []
        for _, row in universe.iterrows():
            code = row["code"]
            ipo = list_date_map.get(code)
            tasks.append((
                code,
                row.get("code_name", ""),
                start_date,
                latest_date,
                date_set,
                cap_map.get(code),
                ipo,
                min_mktcap,
                args.min_list_days,
                args.sleep,
                price_source,
                args.retries,
                args.retry_delay,
                args.debug_filters,
                args.require_end_trade,
                args.missing_mktcap_policy,
                True,
                observation_trade_status.get(code, "unknown"),
            ))

        hits = []
        workers = max(1, args.workers)
        print(
            f"候选股票: {len(tasks)} | offset={args.offset} | limit={args.limit} | "
            f"workers={workers} | price_source={price_source}"
        )
        progress = {
            "processed": 0,
            "succeeded": 0,
            "cache_hits": 0,
            "skipped": 0,
            "failed": 0,
            "signals": 0,
        }
        if workers > 1:
            if bs_available:
                bs.logout()
            initializer = init_worker if price_source == "baostock" else None
            with multiprocessing.Pool(
                processes=workers,
                initializer=initializer,
                maxtasksperchild=args.maxtasksperchild if args.maxtasksperchild > 0 else None,
            ) as pool:
                for idx, result in enumerate(pool.imap_unordered(fetch_one_stock, tasks), start=1):
                    hits.extend(result)
                    update_fetch_progress(progress, result)
                    if idx % 200 == 0 or idx == len(tasks):
                        print(format_fetch_progress(progress, len(tasks)))
                        print("保存方式：每只完成后立即写 DuckDB，并原子更新 CSV；中断后可续跑。")
        else:
            for idx, task in enumerate(tasks, start=1):
                result = fetch_one_stock(task)
                hits.extend(result)
                update_fetch_progress(progress, result)
                if idx % 200 == 0 or idx == len(tasks):
                    print(format_fetch_progress(progress, len(tasks)))
                    print("保存方式：每只完成后立即写 DuckDB，并原子更新 CSV；中断后可续跑。")
            if bs_available:
                bs.logout()

        all_rows_df = pd.DataFrame(hits)
        if all_rows_df.empty:
            statuses_df = pd.DataFrame(
                columns=["code", "name", "latest_data_date", "observation_status", "error"]
            )
            coverage_df = pd.DataFrame(columns=["code", "covered_dates"])
            hits_df = pd.DataFrame()
        else:
            statuses_df = all_rows_df[all_rows_df["signal_type"] == "STATUS"].copy()
            coverage_df = all_rows_df[
                all_rows_df["signal_type"] == "TRADE_COVERAGE"
            ].copy()
            hits_df = all_rows_df[
                ~all_rows_df["signal_type"].isin({"STATUS", "TRADE_COVERAGE"})
            ].copy()
        if args.debug_filters and not hits_df.empty and "signal_type" in hits_df.columns:
            debug_df = hits_df[hits_df["signal_type"] == "DEBUG"].copy()
            if not debug_df.empty:
                latest_debug = debug_df[debug_df["date"] == latest_date]
                print("\n--- 今日33公式分步诊断 ---")
                for col in ["kd80", "wr3", "rsi70", "mktcap_ok", "list_days_ok", "base_ok", "xg_ok"]:
                    if col in latest_debug.columns:
                        print(f"{col}: {int(latest_debug[col].fillna(False).sum())}")
                hits_df = hits_df[hits_df["signal_type"] != "DEBUG"].copy()
        if hits_df.empty:
            window_base_unique = 0
            window_xg_unique = 0
            window_xg_technical_unique = 0
            window_market_cap_unique = 0
            window_market_cap_technical_unique = 0
            suspended_count = 0
            unavailable_count = int(
                statuses_df["observation_status"].eq("data_unavailable").sum()
            ) if not statuses_df.empty else 0
            unique_xg_hits = pd.DataFrame()
            technical_unique_xg_hits = pd.DataFrame()
            market_cap_unique_xg_hits = pd.DataFrame()
            suspended_technical_hits = pd.DataFrame()
        else:
            xg_hits = hits_df[hits_df["signal_type"] == "XG"]
            base_hits = hits_df[hits_df["signal_type"] == "BASE"]
            market_cap_xg_hits = hits_df[
                hits_df["signal_type"] == "MARKET_CAP_XG"
            ]
            formal_window_dates = trade_dates[-21:]
            window_base_unique = int(
                base_hits[base_hits["date"].isin(formal_window_dates)]["code"].nunique()
            )
            window_xg_hits = xg_hits[xg_hits["date"].isin(formal_window_dates)].copy()
            technical_unique_xg_hits, eligible_unique_xg_hits = select_window_unique_hits(
                window_xg_hits,
                statuses_df,
            )
            unique_xg_hits = eligible_unique_xg_hits
            window_xg_unique = int(unique_xg_hits["code"].nunique())
            window_xg_technical_unique = int(technical_unique_xg_hits["code"].nunique())
            window_market_cap_hits = market_cap_xg_hits[
                market_cap_xg_hits["date"].isin(formal_window_dates)
            ].copy()
            (
                market_cap_technical_unique_xg_hits,
                market_cap_tradable_unique_xg_hits,
            ) = select_window_unique_hits(window_market_cap_hits, statuses_df)
            market_cap_unique_xg_hits = market_cap_tradable_unique_xg_hits
            window_market_cap_unique = int(
                market_cap_unique_xg_hits["code"].nunique()
            )
            window_market_cap_technical_unique = int(
                market_cap_technical_unique_xg_hits["code"].nunique()
            )
            status_by_code = (
                statuses_df.drop_duplicates("code", keep="last")
                .set_index("code")["observation_status"]
                if not statuses_df.empty
                else pd.Series(dtype=str)
            )
            suspended_count = int(
                technical_unique_xg_hits["code"]
                .map(status_by_code)
                .eq("suspended_or_no_trade")
                .sum()
            )
            suspended_technical_hits = technical_unique_xg_hits[
                technical_unique_xg_hits["code"]
                .map(status_by_code)
                .eq("suspended_or_no_trade")
            ].reset_index(drop=True)
            unavailable_count = int(status_by_code.eq("data_unavailable").sum())
        summary = build_formula_summary(
            hits_df,
            trade_dates,
            output_days=args.lookback,
            trade_coverage=coverage_df,
            current_statuses=statuses_df,
        )
        if not summary.empty:
            latest_idx = summary.index[-1]
            summary_formal = int(summary.loc[latest_idx, "window_unique_count"])
            summary_technical = int(
                summary.loc[latest_idx, "technical_unique_count"]
            )
            if (
                summary_formal != window_xg_unique
                or summary_technical != window_xg_technical_unique
            ):
                raise RuntimeError(
                    "Formula33 rolling summary is inconsistent with the latest "
                    "formal pool: "
                    f"summary={summary_formal}/{summary_technical} "
                    f"pool={window_xg_unique}/{window_xg_technical_unique}"
                )
            summary.loc[latest_idx, "market_cap_unique_count"] = window_market_cap_unique
            summary.loc[
                latest_idx, "market_cap_technical_unique_count"
            ] = window_market_cap_technical_unique
            summary.loc[latest_idx, "suspended_count"] = suspended_count
            summary.loc[latest_idx, "unavailable_count"] = unavailable_count
        retryable_unavailable = pd.DataFrame()
        if not statuses_df.empty:
            status_errors = statuses_df.get(
                "error",
                pd.Series("", index=statuses_df.index, dtype=str),
            ).fillna("").astype(str).str.strip()
            retryable_unavailable = statuses_df[
                statuses_df["observation_status"].eq("data_unavailable")
                & ~status_errors.eq("insufficient kline history")
            ]
        if not retryable_unavailable.empty:
            codes = ", ".join(
                retryable_unavailable["code"].astype(str).head(10).tolist()
            )
            raise RuntimeError(
                "Formula33 has retryable unavailable stocks; "
                f"completed manifest was not written: count={len(retryable_unavailable)} "
                f"codes={codes}"
            )
        xlsx_path, csv_path = save_workbook(
            summary,
            hits_df,
            unique_hits=unique_xg_hits,
            technical_unique_hits=technical_unique_xg_hits,
            market_cap_unique_hits=market_cap_unique_xg_hits,
            suspended_technical_hits=suspended_technical_hits,
            statuses=statuses_df,
        )
        missing_outputs = [
            path for path in (xlsx_path, csv_path) if not os.path.isfile(path)
        ]
        if missing_outputs:
            raise RuntimeError(
                "Formula33 output write did not complete: "
                + ", ".join(missing_outputs)
            )
        completion_manifest.finish(
            observation_date=latest_date,
            arguments=completion_arguments,
            universe_codes=universe_codes,
            outputs=[xlsx_path, csv_path],
            summary={
                "universe": len(universe_codes),
                "summary_rows": len(summary),
                "market_cap_source": cap_source,
                "list_date_coverage": list_date_coverage,
                "market_cap_coverage": market_cap_coverage,
                "observation_status_coverage": observation_status_coverage,
                "retryable_unavailable_count": 0,
                "window_base_unique": window_base_unique,
                "window_xg_technical_unique": window_xg_technical_unique,
                "window_xg_unique": window_xg_unique,
                "window_market_cap_technical_unique": window_market_cap_technical_unique,
                "window_market_cap_unique": window_market_cap_unique,
                "suspended_count": suspended_count,
                "unavailable_count": unavailable_count,
            },
            code_version=FORMULA33_CODE_VERSION,
        )
        print(f"Excel已保存: {xlsx_path}")
        print(f"CSV已保存: {csv_path}")
        print(summary.to_string(index=False))
        print(f"最近21个交易日BASE去重股票数: {window_base_unique}")
        print(
            "最近21个交易日XG可交易技术去重股票数: "
            f"{window_xg_unique}"
        )
        print(
            f"技术全量: {window_xg_technical_unique} | "
            f"观察日无交易技术命中: {suspended_count} | "
            f"市值大于100亿池: {window_market_cap_unique} | "
            f"数据不可用: {unavailable_count}"
        )
    finally:
        if bs_available:
            try:
                bs.logout()
            except Exception:
                pass
