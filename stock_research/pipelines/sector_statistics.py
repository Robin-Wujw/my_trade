# -*- coding: utf-8 -*-
"""
Sector review statistics in an Excel layout.

The workbook is designed for daily review:
1. Recent dates as columns.
2. Fixed sector groups as rows.
3. Amount leadership, limit-up breadth, wave candidates, and opening patterns.
4. A text-analysis sheet for quick reading.
"""
import argparse
import hashlib
import os
import time
from datetime import datetime

from stock_research.api import akshare as ak
from stock_research.api import ths
from stock_research.api.retry import call_with_backoff
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from stock_research.core.paths import PATHS
from stock_research.core.part_logger import PartLogger
from stock_research.market.sectors import normalize_board_name as normalize_sector_board_name
from stock_research.market.sector_provider import (
    coverage_can_still_pass,
    effective_pipeline_retries,
    sector_history_is_fresh,
    validate_sector_histories,
)
from stock_research.storage import Database, SectorRepository


OUTPUT_DIR = str(PATHS.market_exports)
CACHE_DIR = str(PATHS.cache / "sector_stats")
BENCHMARK_SYMBOL = "sh000001"
BOARD_LIST_MAX_AGE = "24h"
THS_BOARD_SOURCE = "ths/board_list"
THS_HISTORY_SOURCE = "ths/history"
MINIMUM_SECTOR_COVERAGE = 0.95
MINIMUM_STATS_HISTORY_ROWS = 21
MAXIMUM_REQUIRED_HISTORY_ROWS = 60
SECTOR_GROUPS = [
    "有色资源类",
    "半导体",
    "元器件",
    "通信设备",
    "电气设备",
    "其他板块",
]

GROUP_KEYWORDS = {
    "有色资源类": [
        "有色", "金属", "贵金属", "小金属", "能源金属", "稀土", "煤炭", "石油",
        "化工", "化肥", "钢铁", "黄金", "矿",
    ],
    "半导体": ["半导体", "芯片", "集成电路"],
    "元器件": ["元件", "电子元件", "消费电子", "光学光电子", "PCB", "印制电路"],
    "通信设备": ["通信设备", "通信服务", "光通信", "5G"],
    "电气设备": ["电池", "电源设备", "光伏", "风电", "电网", "电机", "电气", "电力设备"],
}


def parse_args(argv=None):
    parser = argparse.ArgumentParser(description="生成类似表格截图的板块复盘统计")
    parser.add_argument("--lookback", type=int, default=10, help="展示最近N个交易日")
    parser.add_argument("--history-days", type=int, default=90, help="板块历史K线拉取天数")
    parser.add_argument("--top-amount", type=int, default=50, help="统计成交额前N板块")
    parser.add_argument("--sleep", type=float, default=0.03, help="板块接口请求间隔")
    parser.add_argument("--retries", type=int, default=4, help="东方财富接口失败重试次数")
    parser.add_argument("--retry-delay", type=float, default=2.0, help="接口失败后的退避基准秒数")
    parser.add_argument("--sample", action="store_true", help="生成离线样例，不访问网络")
    parser.add_argument("--fallback-sample", action="store_true", help="真实板块接口失败时自动生成离线样例，避免每日流程中断")
    return parser.parse_args(argv)


def cache_path(kind, key):
    safe = hashlib.md5(str(key).encode("utf-8")).hexdigest()
    return os.path.join(CACHE_DIR, kind, f"{safe}.csv")


def read_cache(kind, key, date_cols=None):
    path = cache_path(kind, key)
    try:
        if os.path.exists(path):
            df = pd.read_csv(path)
            for col in date_cols or []:
                if col in df.columns:
                    df[col] = pd.to_datetime(df[col], errors="coerce")
            return df
    except Exception as exc:
        print(f"读取缓存失败 {kind}/{key}: {exc}")
    return pd.DataFrame()


def write_cache(kind, key, df):
    if df is None or df.empty:
        return
    path = cache_path(kind, key)
    try:
        os.makedirs(os.path.dirname(path), exist_ok=True)
        tmp_path = f"{path}.{os.getpid()}.tmp"
        df.to_csv(tmp_path, index=False, encoding="utf-8-sig")
        os.replace(tmp_path, path)
    except OSError as exc:
        print(f"写入缓存失败 {kind}/{key}: {exc}")


def classify_group(name):
    text = str(name)
    for group, keywords in GROUP_KEYWORDS.items():
        if any(key in text for key in keywords):
            return group
    return "其他板块"


def normalize_board_name(name):
    return normalize_sector_board_name(name)


def _board_names_from_repository(frame):
    result = frame.rename(
        columns={"board_name": "board", "group_name": "group"}
    ).copy()
    result["group"] = result["group"].fillna("")
    result["group"] = result.apply(
        lambda row: row["group"] or classify_group(row["board"]),
        axis=1,
    )
    if "board_code" not in result.columns:
        result["board_code"] = ""
    result["board_code"] = result["board_code"].fillna("").astype(str)
    return result[["board", "group", "board_code"]].drop_duplicates("board").reset_index(drop=True)


def pct_change(series, days):
    if len(series) <= days:
        return np.nan
    base = series.iloc[-days - 1]
    return series.iloc[-1] / base - 1 if base else np.nan


def candle_label(row):
    pct = row.get("pct_chg", np.nan)
    if pd.isna(pct):
        return "-"
    direction = "阳" if pct >= 0 else "阴"
    apct = abs(pct)
    if apct >= 0.05:
        level = "长"
    elif apct >= 0.02:
        level = "中"
    else:
        level = "小"
    return f"{level}{direction}"


def load_board_names(
    retries=4,
    retry_delay=2.0,
    repository=None,
    logger=None,
    provider=None,
):
    provider = provider or ths
    pipeline_retries = effective_pipeline_retries(provider, retries)
    stale_database = pd.DataFrame()
    if repository is not None:
        try:
            cached = repository.load_boards(
                max_age=BOARD_LIST_MAX_AGE,
                source=THS_BOARD_SOURCE,
            )
            if not cached.empty and cached["board_code"].notna().all():
                result = _board_names_from_repository(cached)
                if logger is not None:
                    logger.event(
                        "board_names",
                        "duckdb",
                        "hit",
                        message="同花顺行业板块列表命中 DuckDB",
                        rows=len(result),
                    )
                return result
            stale_database = repository.load_boards(source=THS_BOARD_SOURCE)
        except Exception as exc:
            if logger is not None:
                logger.event("board_names", "duckdb", "failed", message=str(exc))

    try:
        raw = call_with_backoff(
            provider.load_board_list,
            "同花顺行业板块列表",
            retries=pipeline_retries,
            retry_delay=retry_delay,
        )
    except Exception:
        if stale_database.empty:
            raise
        result = _board_names_from_repository(stale_database)
        result.attrs["offline_cache"] = True
        return result

    if raw is None or raw.empty:
        return pd.DataFrame(columns=["board", "group", "board_code"])
    result = raw.rename(columns={"name": "board", "code": "board_code"}).copy()
    result["board"] = result["board"].map(normalize_board_name)
    result["board_code"] = result["board_code"].astype(str).str.zfill(6)
    result["group"] = result["board"].map(classify_group)
    result = result.dropna(subset=["board"]).drop_duplicates("board").reset_index(drop=True)
    result = result[["board", "group", "board_code"]]
    if repository is not None:
        rows = repository.replace_boards(result, source=THS_BOARD_SOURCE)
        if logger is not None:
            logger.event(
                "board_names",
                "ths",
                "write",
                message="同花顺行业板块快照写入 DuckDB",
                rows=rows,
            )
    write_cache("ths_board_names", "industry", result)
    return result


def load_board_history(
    board,
    days,
    retries=4,
    retry_delay=2.0,
    repository=None,
    logger=None,
    as_of_date=None,
    offline_cache=False,
    board_code=None,
    provider=None,
    request_sleep=0.0,
):
    del offline_cache
    provider = provider or ths
    pipeline_retries = effective_pipeline_retries(provider, retries)
    end_dt = pd.Timestamp(as_of_date).normalize() if as_of_date else pd.Timestamp.today().normalize()
    end_date = end_dt.strftime("%Y%m%d")
    start_date = (end_dt - pd.Timedelta(days=max(days * 3, 180))).strftime("%Y%m%d")
    stored = pd.DataFrame()
    if repository is not None:
        try:
            stored = repository.load_board_history(
                board,
                end_date=end_dt,
                days=days,
                date_column="date",
                source=THS_HISTORY_SOURCE,
            )
            if not stored.empty:
                stored = stored.sort_values("date")
                has_required_depth = len(stored) >= min(days, 60)
                cache_reaches_observation = stored["date"].max() >= end_dt
                if has_required_depth and cache_reaches_observation:
                    if logger is not None:
                        logger.event(
                            "board_history",
                            "duckdb",
                            "hit",
                            message=f"{board} K线命中 DuckDB",
                            rows=len(stored),
                            context={"board": board},
                        )
                    return stored.reset_index(drop=True)
                if has_required_depth:
                    start_date = (
                        stored["date"].max() + pd.Timedelta(days=1)
                    ).strftime("%Y%m%d")
        except Exception as exc:
            if logger is not None:
                logger.event(
                    "board_history",
                    "duckdb",
                    "failed",
                    message=str(exc),
                    context={"board": board},
                )
    if not board_code:
        raise ValueError(f"{board} 缺少同花顺板块代码")
    try:
        if request_sleep > 0:
            time.sleep(request_sleep)
        df = call_with_backoff(
            lambda: provider.load_board_history(
                board_code,
                start_date=start_date,
                end_date=end_date,
            ),
            f"{board} 同花顺板块K线",
            retries=pipeline_retries,
            retry_delay=retry_delay,
        )
        if logger is not None:
            logger.event(
                "board_history",
                "ths",
                "hit",
                message=f"{board} K线使用同花顺",
                rows=len(df),
                context={"board": board, "board_code": str(board_code)},
            )
    except Exception:
        if not stored.empty:
            return stored.reset_index(drop=True)
        raise
    if df is None or df.empty:
        return pd.DataFrame()
    df = df.copy()
    df["date"] = pd.to_datetime(df["date"], errors="coerce")
    df = df.dropna(subset=["date", "close"]).sort_values("date").reset_index(drop=True)
    if repository is not None:
        try:
            rows = repository.upsert_board_history(
                board,
                df,
                source=THS_HISTORY_SOURCE,
            )
            if logger is not None:
                logger.event(
                    "board_history",
                    "duckdb",
                    "write",
                    message=f"{board} K线写入 DuckDB",
                    rows=rows,
                    context={"board": board},
                )
        except Exception as exc:
            if logger is not None:
                logger.event(
                    "board_history",
                    "duckdb",
                    "write_failed",
                    message=str(exc),
                    context={"board": board},
                )
    canonical_columns = [
        "date", "open", "high", "low", "close", "volume", "amount", "pct_chg"
    ]
    combined = pd.concat(
        [
            stored[[column for column in canonical_columns if column in stored.columns]],
            df[[column for column in canonical_columns if column in df.columns]],
        ],
        ignore_index=True,
        sort=False,
    )
    combined = combined.sort_values("date").drop_duplicates("date", keep="last")
    write_cache("ths_board_history", board_code, combined)
    return combined.tail(days).reset_index(drop=True)


def build_board_daily(history_map, board_info, lookback):
    rows = []
    for board, hist in history_map.items():
        if hist.empty:
            continue
        group = board_info.get(board, {}).get("group", classify_group(board))
        for idx in range(len(hist)):
            cur = hist.iloc[: idx + 1]
            row = hist.iloc[idx].to_dict()
            amount = row.get("amount", np.nan)
            amount20 = cur["amount"].tail(20).mean() if "amount" in cur and len(cur) >= 20 else np.nan
            rows.append({
                "date": row["date"],
                "board": board,
                "group": group,
                "close": row.get("close"),
                "open": row.get("open"),
                "amount": amount,
                "pct_chg": row.get("pct_chg"),
                "ret3": pct_change(cur["close"], 3),
                "ret5": pct_change(cur["close"], 5),
                "ret20": pct_change(cur["close"], 20),
                "amount20": amount20,
                "amount_ratio": amount / amount20 if pd.notna(amount) and pd.notna(amount20) and amount20 > 0 else np.nan,
            })
    df = pd.DataFrame(rows)
    if df.empty:
        return df
    df["date_key"] = df["date"].dt.strftime("%Y-%m-%d")
    df["amount_rank"] = df.groupby("date_key")["amount"].rank(ascending=False, method="first")
    df["wave_candidate"] = (
        (df["ret3"].fillna(-9) > 0)
        & (df["ret5"].fillna(-9) > 0.02)
        & (df["ret20"].fillna(-9) > 0)
        & (df["amount_ratio"].fillna(0) >= 1.05)
    )
    dates = sorted(df["date_key"].unique())[-lookback:]
    return df[df["date_key"].isin(dates)].copy()


def load_limit_up_by_date(date_keys):
    frames = []
    for date_key in date_keys:
        cached = read_cache("limit_up_pool", date_key)
        if not cached.empty:
            cached = cached.copy()
            cached["date_key"] = date_key
            cached["code"] = (
                cached["code"].astype(str).str.replace(r"\.0$", "", regex=True).str.zfill(6)
            )
            frames.append(cached)
            continue
        date_str = date_key.replace("-", "")
        try:
            zt = ak.stock_zt_pool_em(date=date_str)
        except Exception as exc:
            raise RuntimeError(f"涨停池 {date_key} 读取失败: {exc}") from exc
        if zt is None or zt.empty:
            raise RuntimeError(f"涨停池 {date_key} 缺失，拒绝按 0 继续统计")
        rows = []
        for _, row in zt.iterrows():
            board = row.get("所属行业") or row.get("行业") or "其他"
            rows.append({
                "date_key": date_key,
                "name": row.get("名称", ""),
                "code": str(row.get("代码", "")).replace(".0", "").zfill(6),
                "board": normalize_board_name(board),
                "group": classify_group(board),
            })
        frame = pd.DataFrame(rows)
        write_cache("limit_up_pool", date_key, frame)
        frames.append(frame)
    if not frames:
        return pd.DataFrame(columns=["date_key", "name", "code", "board", "group"])
    return pd.concat(frames, ignore_index=True)


def _prepare_benchmark_frame(frame):
    required_columns = {"date", "open", "close"}
    if not isinstance(frame, pd.DataFrame) or frame.empty:
        raise RuntimeError("上证指数数据为空")
    missing_columns = required_columns.difference(frame.columns)
    if missing_columns:
        raise RuntimeError(
            f"上证指数缺少必要字段: {sorted(missing_columns)}"
        )
    result = frame.copy()
    result["date"] = pd.to_datetime(result["date"], errors="coerce")
    result["close"] = pd.to_numeric(result["close"], errors="coerce")
    result["open"] = pd.to_numeric(result["open"], errors="coerce")
    result = (
        result.dropna(subset=["date", "open", "close"])
        .sort_values("date")
        .drop_duplicates("date", keep="last")
        .reset_index(drop=True)
    )
    if result.empty:
        raise RuntimeError("上证指数没有可用的日期/开盘/收盘数据")
    return result


def load_benchmark_frame():
    try:
        frame = ak.stock_zh_index_daily(symbol=BENCHMARK_SYMBOL)
    except Exception as exc:
        raise RuntimeError(f"上证指数读取失败: {exc}") from exc
    return _prepare_benchmark_frame(frame)


def load_benchmark_daily(date_keys, benchmark_frame=None):
    df = _prepare_benchmark_frame(
        load_benchmark_frame() if benchmark_frame is None else benchmark_frame
    )
    df["date"] = pd.to_datetime(df["date"], errors="coerce")
    df["date_key"] = df["date"].dt.strftime("%Y-%m-%d")
    df["pct_chg"] = df["close"].pct_change()
    required_dates = list(dict.fromkeys(str(value) for value in date_keys))
    result = df[df["date_key"].isin(required_dates)].copy()
    missing_dates = sorted(set(required_dates).difference(result["date_key"]))
    if missing_dates:
        raise RuntimeError(
            f"上证指数缺少板块统计所需交易日: {', '.join(missing_dates)}"
        )
    return result


def resolve_latest_trading_date(as_of_date=None, benchmark_frame=None):
    cap = pd.Timestamp(as_of_date).normalize() if as_of_date else pd.Timestamp.today().normalize()
    frame = _prepare_benchmark_frame(
        load_benchmark_frame() if benchmark_frame is None else benchmark_frame
    )
    dates = frame["date"]
    dates = dates[dates.notna() & (dates <= cap)]
    if dates.empty:
        raise RuntimeError(f"无法确定 {cap:%Y-%m-%d} 之前的板块观察交易日")
    return dates.max().normalize()


def sample_board_daily(lookback):
    dates = pd.bdate_range(end=pd.Timestamp("2026-06-24"), periods=lookback)
    boards = [
        ("半导体", "半导体"), ("电子元件", "元器件"), ("通信设备", "通信设备"),
        ("电池", "电气设备"), ("能源金属", "有色资源类"), ("软件服务", "其他板块"),
        ("光伏设备", "电气设备"), ("贵金属", "有色资源类"), ("消费电子", "元器件"),
    ]
    rows = []
    for b_idx, (board, group) in enumerate(boards):
        close = 100 + b_idx * 3
        for d_idx, day in enumerate(dates):
            pct = ((b_idx * 7 + d_idx * 5) % 13 - 4) / 100
            close *= 1 + pct
            amount = 80 + ((b_idx + 2) * (d_idx + 3) * 17) % 120
            rows.append({
                "date": day,
                "date_key": day.strftime("%Y-%m-%d"),
                "board": board,
                "group": group,
                "close": close,
                "open": close / (1 + pct * 0.5),
                "amount": amount,
                "pct_chg": pct,
                "ret3": 0.02 + pct,
                "ret5": 0.04 + pct,
                "ret20": 0.12 + pct,
                "amount20": 100,
                "amount_ratio": amount / 100,
            })
    df = pd.DataFrame(rows)
    df["amount_rank"] = df.groupby("date_key")["amount"].rank(ascending=False, method="first")
    df["wave_candidate"] = (df["ret5"] > 0.03) & (df["amount_ratio"] >= 1.05)
    return df


def sample_limit_up(date_keys):
    rows = []
    for idx, date_key in enumerate(date_keys):
        for group in SECTOR_GROUPS[:-1]:
            count = (idx + len(group)) % 4
            for num in range(count):
                rows.append({
                    "date_key": date_key,
                    "name": f"{group[:2]}样本{num + 1}",
                    "code": f"sample{idx}{num}",
                    "board": group,
                    "group": group,
                })
        if idx % 2 == 0:
            rows.append({"date_key": date_key, "name": "软件样本", "code": "samplex", "board": "软件服务", "group": "其他板块"})
    return pd.DataFrame(rows)


def make_count_matrix(df, date_keys, value_col):
    matrix = pd.DataFrame(index=SECTOR_GROUPS, columns=date_keys)
    for date_key in date_keys:
        cur = df[df["date_key"] == date_key]
        counts = cur.groupby("group")[value_col].sum() if value_col in cur else cur.groupby("group").size()
        for group in SECTOR_GROUPS:
            matrix.loc[group, date_key] = int(counts.get(group, 0))
    matrix.loc["合计", :] = matrix.fillna(0).astype(int).sum(axis=0)
    return matrix.fillna("")


def other_detail(df, date_key):
    cur = df[(df["date_key"] == date_key) & (df["group"] == "其他板块")]
    if cur.empty:
        return ""
    counts = cur.groupby("board").size().sort_values(ascending=False).head(3)
    return "，".join(f"{int(v)}（{k}）" for k, v in counts.items())


def build_sections(board_daily, limit_up, top_amount):
    date_keys = sorted(board_daily["date_key"].unique())
    amount_top = board_daily[board_daily["amount_rank"] <= top_amount].copy()
    amount_top["hit"] = 1
    amount_matrix = make_count_matrix(amount_top, date_keys, "hit")

    if limit_up.empty:
        limit_up = pd.DataFrame(columns=["date_key", "group", "board"])
    limit_up["hit"] = 1
    limit_matrix = make_count_matrix(limit_up, date_keys, "hit")
    if "其他板块" in limit_matrix.index:
        for date_key in date_keys:
            limit_matrix.loc["其他板块", date_key] = other_detail(limit_up, date_key)

    wave = board_daily[board_daily["wave_candidate"]].copy()
    wave["hit"] = 1
    wave_matrix = make_count_matrix(wave, date_keys, "hit")

    return date_keys, amount_matrix, limit_matrix, wave_matrix


def build_open_patterns(board_daily, benchmark, date_keys):
    rows = []
    for date_key in date_keys:
        bench_label = "-"
        if benchmark is not None and not benchmark.empty:
            cur_bench = benchmark[benchmark["date_key"] == date_key]
            if not cur_bench.empty:
                bench_label = candle_label(cur_bench.iloc[-1])
        cur = board_daily[board_daily["date_key"] == date_key].copy()
        strong = cur.sort_values(["pct_chg", "amount_ratio"], ascending=False).head(2)
        shrink = cur[(cur["pct_chg"] > 0) & (cur["amount_ratio"] < 1)].sort_values("pct_chg", ascending=False).head(1)
        rows.append({
            "date_key": date_key,
            "大盘K线": bench_label,
            "强势板块": "、".join(strong["board"].tolist()),
            "缩量上涨": "、".join(shrink["board"].tolist()) if not shrink.empty else "",
        })
    return pd.DataFrame(rows)


def write_matrix(ws, start_row, title, matrix, latest_col_fill):
    thin = Side(style="thin", color="666666")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    title_fill = PatternFill("solid", fgColor="E2F0D9")
    header_fill = PatternFill("solid", fgColor="D9EAF7")

    ws.cell(start_row, 1, title)
    ws.cell(start_row, 1).font = Font(bold=True)
    ws.cell(start_row, 1).fill = title_fill
    row = start_row + 1
    ws.cell(row, 1, "分类")
    for col_idx, date_key in enumerate(matrix.columns, start=2):
        ws.cell(row, col_idx, date_key[5:].replace("-", "."))
    for cell in ws[row]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for r_idx, (idx, values) in enumerate(matrix.iterrows(), start=row + 1):
        ws.cell(r_idx, 1, idx)
        for c_idx, value in enumerate(values.tolist(), start=2):
            ws.cell(r_idx, c_idx, value)
    end_row = row + len(matrix)
    latest_col = len(matrix.columns) + 1
    for rows in ws.iter_rows(min_row=row + 1, max_row=end_row, min_col=1, max_col=latest_col):
        for cell in rows:
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            if cell.column == latest_col:
                cell.fill = latest_col_fill
    return end_row + 2


def build_analysis_lines(board_daily, limit_up, date_keys):
    latest = date_keys[-1]
    cur = board_daily[board_daily["date_key"] == latest].copy()
    amount_top = cur.sort_values("amount", ascending=False).head(8)
    wave_top = cur[cur["wave_candidate"]].sort_values(["ret5", "amount_ratio"], ascending=False).head(6)
    limit_cur = limit_up[limit_up["date_key"] == latest] if not limit_up.empty else pd.DataFrame()
    limit_groups = limit_cur.groupby("group").size().sort_values(ascending=False) if not limit_cur.empty else pd.Series(dtype=int)

    lines = [f"一）板块复盘分析（{latest}）", ""]
    for idx, group in enumerate(SECTOR_GROUPS[:-1], start=1):
        boards = amount_top[amount_top["group"] == group]["board"].head(3).tolist()
        waves = wave_top[wave_top["group"] == group]["board"].head(3).tolist()
        limit_count = int(limit_groups.get(group, 0))
        if boards or waves or limit_count:
            lines.append(
                f"{idx}. {group}：成交额靠前板块：{('、'.join(boards) if boards else '无')}；"
                f"涨停扩散数：{limit_count}；三浪候选：{('、'.join(waves) if waves else '无')}。"
            )
    other = limit_groups.get("其他板块", 0)
    if other:
        lines.append(f"{len(SECTOR_GROUPS)}. 其他板块：涨停扩散数 {int(other)}，需要看是否只是轮动补涨。")
    lines.append("")
    lines.append("结论：优先观察连续进入成交额前列、同时涨停扩散和三浪候选都增加的方向；只有价值回归但无量能的板块，放入观察不追。")
    return lines


def save_outputs(board_daily, limit_up, benchmark, top_amount):
    date_keys, amount_matrix, limit_matrix, wave_matrix = build_sections(board_daily, limit_up, top_amount)
    open_patterns = build_open_patterns(board_daily, benchmark, date_keys)
    analysis_lines = build_analysis_lines(board_daily, limit_up, date_keys)

    os.makedirs(OUTPUT_DIR, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    xlsx_path = os.path.join(OUTPUT_DIR, f"sector_stats_{stamp}.xlsx")
    md_path = os.path.join(OUTPUT_DIR, f"sector_stats_{stamp}.md")

    wb = Workbook()
    ws = wb.active
    ws.title = "板块横向统计"
    ws.freeze_panes = "B2"
    latest_fill = PatternFill("solid", fgColor="FFF2CC")
    next_row = 1
    next_row = write_matrix(ws, next_row, "一、成交额Top板块数量", amount_matrix, latest_fill)
    next_row = write_matrix(ws, next_row, "二、涨停扩散数量", limit_matrix, latest_fill)
    next_row = write_matrix(ws, next_row, "三、三浪/放量候选数量", wave_matrix, latest_fill)

    ws.cell(next_row, 1, "四、开盘八法/盘面状态")
    ws.cell(next_row, 1).font = Font(bold=True)
    headers = ["日期", "大盘K线", "强势板块", "缩量上涨"]
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(next_row + 1, col_idx, header)
        ws.cell(next_row + 1, col_idx).font = Font(bold=True)
    for r_idx, row in enumerate(open_patterns.to_dict("records"), start=next_row + 2):
        ws.cell(r_idx, 1, row["date_key"])
        ws.cell(r_idx, 2, row["大盘K线"])
        ws.cell(r_idx, 3, row["强势板块"])
        ws.cell(r_idx, 4, row["缩量上涨"])

    ws2 = wb.create_sheet("文字分析")
    for idx, line in enumerate(analysis_lines, start=1):
        ws2.cell(idx, 1, line)
        ws2.cell(idx, 1).alignment = Alignment(wrap_text=True, vertical="top")
        if idx == 1:
            ws2.cell(idx, 1).font = Font(bold=True, size=13)

    for sheet in wb.worksheets:
        for col in range(1, sheet.max_column + 1):
            sheet.column_dimensions[get_column_letter(col)].width = 16 if col > 1 else 18
        for row in range(1, sheet.max_row + 1):
            sheet.row_dimensions[row].height = 24
    ws2.column_dimensions["A"].width = 120
    wb.save(xlsx_path)

    with open(md_path, "w", encoding="utf-8") as fh:
        fh.write("\n".join(analysis_lines))
        fh.write("\n")
    return xlsx_path, md_path, analysis_lines


def main(argv=None, repository=None, logger=None, provider=None):
    args = parse_args(argv)
    provider = provider or ths
    if repository is None:
        database = Database()
        database.initialize()
        repository = SectorRepository(database)
    if logger is None:
        logger = PartLogger("sector_stats", repository=repository)
    if args.sample:
        board_daily = sample_board_daily(args.lookback)
        date_keys = sorted(board_daily["date_key"].unique())
        limit_up = sample_limit_up(date_keys)
        benchmark = pd.DataFrame({
            "date_key": date_keys,
            "pct_chg": [0.01, -0.008, 0.018, 0.026, -0.004, 0.012, 0.006, 0.021, 0.009, 0.015][-len(date_keys):],
        })
    else:
        benchmark_frame = load_benchmark_frame()
        observation_date = resolve_latest_trading_date(
            benchmark_frame=benchmark_frame
        )
        try:
            with logger.part("board_names"):
                boards = load_board_names(
                    retries=args.retries,
                    retry_delay=args.retry_delay,
                    repository=repository,
                    logger=logger,
                    provider=provider,
                )
        except Exception as exc:
            if args.fallback_sample:
                print(f"真实板块接口失败，改用 --fallback-sample 样例数据: {exc}")
                board_daily = sample_board_daily(args.lookback)
                date_keys = sorted(board_daily["date_key"].unique())
                limit_up = sample_limit_up(date_keys)
                benchmark = pd.DataFrame({
                    "date_key": date_keys,
                    "pct_chg": [0.01, -0.008, 0.018, 0.026, -0.004, 0.012, 0.006, 0.021, 0.009, 0.015][-len(date_keys):],
                })
                xlsx_path, md_path, analysis_lines = save_outputs(board_daily, limit_up, benchmark, args.top_amount)
                print(f"Excel已保存: {xlsx_path}")
                print(f"文字分析已保存: {md_path}")
                print("\n".join(analysis_lines[:8]))
                return
            raise SystemExit(f"无法访问同花顺板块接口：{exc}\n可先用 --sample 查看版式，或换网络后重跑真实数据。")
        if boards.empty:
            if args.fallback_sample:
                print("真实板块列表为空，改用 --fallback-sample 样例数据")
                board_daily = sample_board_daily(args.lookback)
                date_keys = sorted(board_daily["date_key"].unique())
                limit_up = sample_limit_up(date_keys)
                benchmark = pd.DataFrame({"date_key": date_keys, "pct_chg": [0.0] * len(date_keys)})
                xlsx_path, md_path, analysis_lines = save_outputs(board_daily, limit_up, benchmark, args.top_amount)
                print(f"Excel已保存: {xlsx_path}")
                print(f"文字分析已保存: {md_path}")
                print("\n".join(analysis_lines[:8]))
                return
            raise SystemExit("无法获取行业板块列表")
        board_info = boards.set_index("board").to_dict("index")
        history_map = {}
        offline_cache = bool(boards.attrs.get("offline_cache"))
        expected_boards = len(boards)
        fresh_completed = 0
        history_days = max(MINIMUM_STATS_HISTORY_ROWS, args.history_days)
        minimum_history_rows = min(history_days, MAXIMUM_REQUIRED_HISTORY_ROWS)
        with logger.part("board_history"):
            for completed, (_, board_row) in enumerate(boards.iterrows(), start=1):
                board = board_row["board"]
                try:
                    history = load_board_history(
                        board,
                        history_days,
                        retries=args.retries,
                        retry_delay=args.retry_delay,
                        repository=repository,
                        logger=logger,
                        offline_cache=offline_cache,
                        as_of_date=observation_date,
                        board_code=board_row.get("board_code"),
                        provider=provider,
                        request_sleep=args.sleep,
                    )
                    history_map[board] = history
                    if sector_history_is_fresh(
                        history,
                        observation_date=observation_date,
                        max_stale_days=0,
                        minimum_rows=minimum_history_rows,
                    ):
                        fresh_completed += 1
                except Exception as exc:
                    print(f"跳过 {board}: {exc}")
                    logger.event(
                        "board_history",
                        "board",
                        "skip",
                        message=str(exc),
                        context={"board": board},
                    )
                if not coverage_can_still_pass(
                    expected=expected_boards,
                    completed=completed,
                    fresh=fresh_completed,
                    minimum=MINIMUM_SECTOR_COVERAGE,
                ):
                    message = (
                        f"提前停止: completed={completed} fresh={fresh_completed} "
                        f"remaining={expected_boards - completed}，已无法达到 "
                        f"{MINIMUM_SECTOR_COVERAGE:.0%} 覆盖率"
                    )
                    print(message)
                    logger.event(
                        "board_history",
                        "early_stop",
                        "failed",
                        message=message,
                        rows=fresh_completed,
                    )
                    break
                if completed % 20 == 0:
                    print(f"板块K线进度 {completed}/{len(boards)}")
                    logger.event(
                        "board_history",
                        "progress",
                        "running",
                        message=f"板块K线进度 {completed}/{len(boards)}",
                        rows=len(history_map),
                    )
        history_map, coverage = validate_sector_histories(
            boards["board"].tolist(),
            history_map,
            observation_date=observation_date,
            max_stale_days=0,
            minimum_rows=minimum_history_rows,
            minimum=MINIMUM_SECTOR_COVERAGE,
        )
        coverage_status = "pass" if coverage.passed else "failed"
        coverage_message = (
            f"provider=ths expected={coverage.expected} fresh={coverage.fresh} "
            f"stale={coverage.stale} missing={coverage.missing} "
            f"coverage={coverage.coverage:.1%}"
        )
        print(f"板块覆盖率: {coverage_message}")
        logger.event(
            "coverage",
            "ths",
            coverage_status,
            message=coverage_message,
            rows=coverage.fresh,
            context={
                "expected": coverage.expected,
                "fresh": coverage.fresh,
                "stale": coverage.stale,
                "missing": coverage.missing,
                "coverage": coverage.coverage,
            },
        )
        if not coverage.passed:
            raise SystemExit(2)
        board_daily = build_board_daily(history_map, board_info, args.lookback)
        if board_daily.empty:
            if args.fallback_sample:
                print("真实板块历史为空，改用 --fallback-sample 样例数据")
                board_daily = sample_board_daily(args.lookback)
                date_keys = sorted(board_daily["date_key"].unique())
                limit_up = sample_limit_up(date_keys)
                benchmark = pd.DataFrame({"date_key": date_keys, "pct_chg": [0.0] * len(date_keys)})
                xlsx_path, md_path, analysis_lines = save_outputs(board_daily, limit_up, benchmark, args.top_amount)
                print(f"Excel已保存: {xlsx_path}")
                print(f"文字分析已保存: {md_path}")
                print("\n".join(analysis_lines[:8]))
                return
            raise SystemExit("无有效板块历史数据")
        date_keys = sorted(board_daily["date_key"].unique())
        limit_up = load_limit_up_by_date(date_keys)
        benchmark = load_benchmark_daily(
            date_keys,
            benchmark_frame=benchmark_frame,
        )

    xlsx_path, md_path, analysis_lines = save_outputs(board_daily, limit_up, benchmark, args.top_amount)
    print(f"Excel已保存: {xlsx_path}")
    print(f"文字分析已保存: {md_path}")
    print("\n".join(analysis_lines[:8]))
