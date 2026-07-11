"""Verify a Formula33 workbook against the reviewed screenshot baseline."""
from __future__ import annotations

import argparse
import json
import re
from pathlib import Path

from openpyxl import load_workbook


SHEETS = {
    "formal_count": "21日技术可交易",
    "technical_count": "21日技术全量",
    "market_cap_count": "市值大于100亿池",
    "suspended_count": "停牌技术命中诊断",
}


def _normalize_code(value):
    match = re.search(r"(\d{6})$", str(value or "").strip())
    return match.group(1) if match else ""


def _sheet_codes(workbook, sheet_name):
    sheet = workbook[sheet_name]
    rows = sheet.iter_rows(values_only=True)
    headers = [str(value or "").strip() for value in next(rows)]
    try:
        code_index = headers.index("code")
    except ValueError as exc:
        raise AssertionError(f"{sheet_name} is missing the code column") from exc
    return {
        code
        for row in rows
        if code_index < len(row) and (code := _normalize_code(row[code_index]))
    }


def verify_workbook(workbook_path, expected_path):
    expected = json.loads(Path(expected_path).read_text(encoding="utf-8"))
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        formal_codes = _sheet_codes(workbook, SHEETS["formal_count"])
        expected_codes = set(expected["codes"])
        missing = sorted(expected_codes - formal_codes)
        extra = sorted(formal_codes - expected_codes)
        if missing or extra:
            raise AssertionError(
                f"Formula33 formal codes differ: missing={missing} extra={extra}"
            )

        counts = {"formal_count": len(formal_codes)}
        for key, sheet_name in SHEETS.items():
            if key != "formal_count":
                counts[key] = len(_sheet_codes(workbook, sheet_name))
        for key, actual in counts.items():
            wanted = int(expected[key])
            if actual != wanted:
                raise AssertionError(
                    f"Formula33 {key} differs: expected={wanted} actual={actual}"
                )

        summary = workbook["33公式日统计"]
        rows = list(summary.iter_rows(values_only=True))
        headers = [str(value or "").strip() for value in rows[0]]
        records = [dict(zip(headers, row)) for row in rows[1:] if row[0]]
        actual_start = str(records[0]["date"])[:10]
        actual_end = str(records[-1]["date"])[:10]
        if (actual_start, actual_end) != (
            expected["start_date"],
            expected["end_date"],
        ):
            raise AssertionError(
                "Formula33 date range differs: "
                f"expected={expected['start_date']}..{expected['end_date']} "
                f"actual={actual_start}..{actual_end}"
            )
        if int(records[-1]["window_unique_count"]) != int(expected["formal_count"]):
            raise AssertionError("latest summary count does not match formal sheet")
        return counts
    finally:
        workbook.close()


def main(argv=None):
    parser = argparse.ArgumentParser(
        description="Verify a Formula33 workbook against a screenshot baseline"
    )
    parser.add_argument("workbook")
    parser.add_argument("expected")
    args = parser.parse_args(argv)
    counts = verify_workbook(args.workbook, args.expected)
    print(
        "Formula33 screenshot verified: "
        f"formal={counts['formal_count']} technical={counts['technical_count']} "
        f"suspended={counts['suspended_count']} market_cap={counts['market_cap_count']}"
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
